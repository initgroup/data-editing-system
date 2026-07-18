"""Build the M90003 initial column-type training profile pack.

Only derived column statistics and source attribution are written to the
repository. Raw public datasets remain in the maintainer's temporary cache.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import random
import re
import urllib.request
import zipfile
from collections import Counter
from pathlib import Path
from typing import Iterable, Iterator


SAMPLE_LIMIT = 10_000
MISSING_VALUES = {"", "?", "NA", "N/A", "NULL", "NONE", "NAN"}
TYPE_GROUPS = {
    "NUM_IDENTIFIER": "OTHER",
    "CHAR_IDENTIFIER": "OTHER",
    "NUM_CONTINUOUS": "CONTINUOUS",
    "NUM_DISCRETE": "CONTINUOUS",
    "CAT_GENERAL": "CATEGORICAL",
    "CAT_CHAR": "CATEGORICAL",
    "CAT_ORDINAL": "CATEGORICAL",
    "CAT_NUMERIC": "CATEGORICAL",
    "FREE_TEXT": "OTHER",
    "OTHER": "OTHER",
}

SOURCES = {
    "adult": {
        "datasetId": 2,
        "title": "Adult",
        "doi": "10.24432/C5XW20",
        "landingPage": "https://archive.ics.uci.edu/dataset/2/adult",
        "url": "https://archive.ics.uci.edu/static/public/2/adult.zip",
    },
    "bank": {
        "datasetId": 222,
        "title": "Bank Marketing",
        "doi": "10.24432/C5K306",
        "landingPage": "https://archive.ics.uci.edu/dataset/222/bank+marketing",
        "url": "https://archive.ics.uci.edu/static/public/222/bank+marketing.zip",
    },
    "bike": {
        "datasetId": 275,
        "title": "Bike Sharing",
        "doi": "10.24432/C5W894",
        "landingPage": "https://archive.ics.uci.edu/dataset/275/bike+sharing+dataset",
        "url": "https://archive.ics.uci.edu/static/public/275/bike+sharing+dataset.zip",
    },
    "wine": {
        "datasetId": 186,
        "title": "Wine Quality",
        "doi": "10.24432/C56S3T",
        "landingPage": "https://archive.ics.uci.edu/dataset/186/wine+quality",
        "url": "https://archive.ics.uci.edu/static/public/186/wine+quality.zip",
    },
    "sms": {
        "datasetId": 228,
        "title": "SMS Spam Collection",
        "doi": "10.24432/C5CC84",
        "landingPage": "https://archive.ics.uci.edu/dataset/228/sms+spam+collection",
        "url": "https://archive.ics.uci.edu/static/public/228/sms+spam+collection.zip",
    },
    "drybean": {
        "datasetId": 602,
        "title": "Dry Bean",
        "doi": "10.24432/C50S4B",
        "landingPage": "https://archive.ics.uci.edu/dataset/602/dry+bean+dataset",
        "url": "https://archive.ics.uci.edu/static/public/602/dry+bean+dataset.zip",
    },
    "seoulbike": {
        "datasetId": 560,
        "title": "Seoul Bike Sharing Demand",
        "doi": "10.24432/C5F62R",
        "landingPage": "https://archive.ics.uci.edu/dataset/560/seoul+bike+sharing+demand",
        "url": "https://archive.ics.uci.edu/static/public/560/seoul+bike+sharing+demand.zip",
    },
    "retail": {
        "datasetId": 352,
        "title": "Online Retail",
        "doi": "10.24432/C5BW33",
        "landingPage": "https://archive.ics.uci.edu/dataset/352/online+retail",
        "url": "https://archive.ics.uci.edu/static/public/352/online+retail.zip",
    },
}

for source in SOURCES.values():
    source["license"] = "CC BY 4.0"
    source["repository"] = "UCI Machine Learning Repository"


def field(type_code: str, data_type: str, description: str) -> dict[str, str]:
    return {"typeCode": type_code, "dataType": data_type, "description": description}


ADULT_COLUMNS = [
    "age", "workclass", "fnlwgt", "education", "education_num", "marital_status",
    "occupation", "relationship", "race", "sex", "capital_gain", "capital_loss",
    "hours_per_week", "native_country", "income",
]
ADULT_LABELS = {
    "age": field("NUM_CONTINUOUS", "NUMBER", "Age"),
    "workclass": field("CAT_GENERAL", "VARCHAR2", "Employment class"),
    "fnlwgt": field("NUM_CONTINUOUS", "NUMBER", "Census final weight"),
    "education": field("CAT_ORDINAL", "VARCHAR2", "Education level"),
    "education_num": field("CAT_ORDINAL", "NUMBER", "Encoded education level"),
    "marital_status": field("CAT_GENERAL", "VARCHAR2", "Marital status"),
    "occupation": field("CAT_GENERAL", "VARCHAR2", "Occupation"),
    "relationship": field("CAT_GENERAL", "VARCHAR2", "Household relationship"),
    "race": field("CAT_GENERAL", "VARCHAR2", "Race category"),
    "sex": field("CAT_CHAR", "CHAR", "Sex category"),
    "capital_gain": field("NUM_CONTINUOUS", "NUMBER", "Capital gain amount"),
    "capital_loss": field("NUM_CONTINUOUS", "NUMBER", "Capital loss amount"),
    "hours_per_week": field("NUM_CONTINUOUS", "NUMBER", "Hours worked per week"),
    "native_country": field("CAT_GENERAL", "VARCHAR2", "Native country"),
    "income": field("CAT_ORDINAL", "VARCHAR2", "Income band"),
}

BANK_LABELS = {
    "age": field("NUM_CONTINUOUS", "NUMBER", "Customer age"),
    "job": field("CAT_GENERAL", "VARCHAR2", "Job category"),
    "marital": field("CAT_GENERAL", "VARCHAR2", "Marital status"),
    "education": field("CAT_ORDINAL", "VARCHAR2", "Education level"),
    "default": field("CAT_CHAR", "CHAR", "Credit default flag"),
    "balance": field("NUM_CONTINUOUS", "NUMBER", "Average yearly balance"),
    "housing": field("CAT_CHAR", "CHAR", "Housing loan flag"),
    "loan": field("CAT_CHAR", "CHAR", "Personal loan flag"),
    "contact": field("CAT_GENERAL", "VARCHAR2", "Contact channel"),
    "day": field("NUM_DISCRETE", "NUMBER", "Contact day of month"),
    "month": field("CAT_ORDINAL", "VARCHAR2", "Contact month"),
    "duration": field("NUM_CONTINUOUS", "NUMBER", "Contact duration"),
    "campaign": field("NUM_DISCRETE", "NUMBER", "Campaign contact count"),
    "pdays": field("NUM_DISCRETE", "NUMBER", "Days since previous contact"),
    "previous": field("NUM_DISCRETE", "NUMBER", "Previous contact count"),
    "poutcome": field("CAT_GENERAL", "VARCHAR2", "Previous campaign outcome"),
    "y": field("CAT_CHAR", "CHAR", "Term deposit response"),
}

BIKE_LABELS = {
    "instant": field("NUM_IDENTIFIER", "NUMBER", "Row identifier"),
    "dteday": field("OTHER", "DATE", "Observation date"),
    "season": field("CAT_ORDINAL", "NUMBER", "Season code"),
    "yr": field("CAT_ORDINAL", "NUMBER", "Year code"),
    "mnth": field("CAT_ORDINAL", "NUMBER", "Month code"),
    "hr": field("CAT_ORDINAL", "NUMBER", "Hour code"),
    "holiday": field("CAT_NUMERIC", "NUMBER", "Holiday flag"),
    "weekday": field("CAT_ORDINAL", "NUMBER", "Weekday code"),
    "workingday": field("CAT_NUMERIC", "NUMBER", "Working day flag"),
    "weathersit": field("CAT_ORDINAL", "NUMBER", "Weather situation code"),
    "temp": field("NUM_CONTINUOUS", "NUMBER", "Normalized temperature"),
    "atemp": field("NUM_CONTINUOUS", "NUMBER", "Normalized apparent temperature"),
    "hum": field("NUM_CONTINUOUS", "NUMBER", "Normalized humidity"),
    "windspeed": field("NUM_CONTINUOUS", "NUMBER", "Normalized wind speed"),
    "casual": field("NUM_CONTINUOUS", "NUMBER", "Casual rental count"),
    "registered": field("NUM_CONTINUOUS", "NUMBER", "Registered rental count"),
    "cnt": field("NUM_CONTINUOUS", "NUMBER", "Total rental count"),
}

WINE_LABELS = {
    "fixed acidity": field("NUM_CONTINUOUS", "NUMBER", "Fixed acidity"),
    "volatile acidity": field("NUM_CONTINUOUS", "NUMBER", "Volatile acidity"),
    "citric acid": field("NUM_CONTINUOUS", "NUMBER", "Citric acid"),
    "residual sugar": field("NUM_CONTINUOUS", "NUMBER", "Residual sugar"),
    "chlorides": field("NUM_CONTINUOUS", "NUMBER", "Chlorides"),
    "free sulfur dioxide": field("NUM_CONTINUOUS", "NUMBER", "Free sulfur dioxide"),
    "total sulfur dioxide": field("NUM_CONTINUOUS", "NUMBER", "Total sulfur dioxide"),
    "density": field("NUM_CONTINUOUS", "NUMBER", "Density"),
    "pH": field("NUM_CONTINUOUS", "NUMBER", "pH"),
    "sulphates": field("NUM_CONTINUOUS", "NUMBER", "Sulphates"),
    "alcohol": field("NUM_CONTINUOUS", "NUMBER", "Alcohol"),
    "quality": field("CAT_ORDINAL", "NUMBER", "Wine quality score"),
}

SEOUL_LABELS = {
    "Date": field("OTHER", "DATE", "Rental date"),
    "Rented Bike Count": field("NUM_CONTINUOUS", "NUMBER", "Hourly rental count"),
    "Hour": field("CAT_ORDINAL", "NUMBER", "Hour code"),
    "Temperature(°C)": field("NUM_CONTINUOUS", "NUMBER", "Temperature"),
    "Humidity(%)": field("NUM_CONTINUOUS", "NUMBER", "Humidity"),
    "Wind speed (m/s)": field("NUM_CONTINUOUS", "NUMBER", "Wind speed"),
    "Visibility (10m)": field("NUM_CONTINUOUS", "NUMBER", "Visibility"),
    "Dew point temperature(°C)": field("NUM_CONTINUOUS", "NUMBER", "Dew point temperature"),
    "Solar Radiation (MJ/m2)": field("NUM_CONTINUOUS", "NUMBER", "Solar radiation"),
    "Rainfall(mm)": field("NUM_CONTINUOUS", "NUMBER", "Rainfall"),
    "Snowfall (cm)": field("NUM_CONTINUOUS", "NUMBER", "Snowfall"),
    "Seasons": field("CAT_ORDINAL", "VARCHAR2", "Season"),
    "Holiday": field("CAT_CHAR", "VARCHAR2", "Holiday flag"),
    "Functioning Day": field("CAT_CHAR", "VARCHAR2", "Service functioning flag"),
}

RETAIL_LABELS = {
    "InvoiceNo": field("CHAR_IDENTIFIER", "VARCHAR2", "Invoice identifier"),
    "StockCode": field("CHAR_IDENTIFIER", "VARCHAR2", "Product identifier"),
    "Description": field("FREE_TEXT", "VARCHAR2", "Product description"),
    "Quantity": field("NUM_CONTINUOUS", "NUMBER", "Transaction quantity"),
    "InvoiceDate": field("OTHER", "TIMESTAMP", "Invoice timestamp"),
    "UnitPrice": field("NUM_CONTINUOUS", "NUMBER", "Unit price"),
    "CustomerID": field("NUM_IDENTIFIER", "NUMBER", "Customer identifier"),
    "Country": field("CAT_GENERAL", "VARCHAR2", "Customer country"),
}


def normalize(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return None if text.upper() in MISSING_VALUES else text


def oracle_number(value: str | None) -> float | None:
    if value is None or not re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?", value):
        return None
    unsigned = value.lstrip("+-")
    if re.match(r"^0\d", unsigned) and not re.match(r"^0$|^0\.", unsigned):
        return None
    try:
        number = float(value)
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def safe_name(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_$#]+", "_", value.strip()).strip("_").upper()
    return (text or "COLUMN")[:128]


def stream_profile_rows(
    rows: Iterable[list[object]],
    columns: list[str],
    labels: dict[str, dict[str, str]],
    source_key: str,
    table_name: str,
) -> list[dict[str, object]]:
    total_rows = 0
    full_non_null = [0] * len(columns)
    full_distinct = [set() for _ in columns]
    reservoir: list[list[str | None]] = []
    rng = random.Random(42)

    for raw_row in rows:
        if len(raw_row) < len(columns):
            continue
        row = [normalize(raw_row[index]) for index in range(len(columns))]
        total_rows += 1
        for index, value in enumerate(row):
            if value is not None:
                full_non_null[index] += 1
                full_distinct[index].add(value)
        if len(reservoir) < SAMPLE_LIMIT:
            reservoir.append(row)
        else:
            slot = rng.randrange(total_rows)
            if slot < SAMPLE_LIMIT:
                reservoir[slot] = row

    source = SOURCES[source_key]
    profiles: list[dict[str, object]] = []
    for index, column in enumerate(columns):
        label = labels.get(column)
        if not label:
            continue
        sample_values = [row[index] for row in reservoir]
        values = [value for value in sample_values if value is not None]
        frequencies = Counter(values)
        distinct_count = len(full_distinct[index])
        sample_distinct = len(frequencies)
        sample_not_null = len(values)
        entropy = 0.0
        if sample_not_null:
            entropy = -sum(
                (count / sample_not_null) * math.log(count / sample_not_null)
                for count in frequencies.values()
            )
        norm_entropy = entropy / math.log(sample_distinct) if sample_distinct > 1 else 0.0
        numbers = [number for number in (oracle_number(value) for value in values) if number is not None]
        integer_count = sum(
            1 for number in numbers
            if abs(number - round(number)) <= 1e-9 * max(1.0, abs(number))
        )
        data_type = label["dataType"]
        numeric_ratio = len(numbers) / sample_not_null if sample_not_null else 0.0
        log_data_type = (
            "NUM" if data_type in {"NUMBER", "FLOAT", "BINARY_FLOAT", "BINARY_DOUBLE"}
            else "ETC" if not sample_not_null
            else "NUM" if numeric_ratio >= 0.98
            else "CHR"
        )
        profile = {
            "runSourceType": "DATA_WORK",
            "runId": 0,
            "owner": "INIT$SAMPLE",
            "tableName": safe_name(f"UCI_{source['datasetId']:04d}_{table_name}"),
            "columnName": safe_name(column),
            "featureVersion": "V2",
            "columnDesc": label["description"],
            "columnId": index + 1,
            "dataType": data_type,
            "totalRows": total_rows,
            "nonNullRows": full_non_null[index],
            "sampleRows": len(reservoir),
            "sampleNotNullRows": sample_not_null,
            "numDistinct": distinct_count,
            "sampleDistinct": sample_distinct,
            "distinctRatio": round(distinct_count / full_non_null[index], 6) if full_non_null[index] else 0,
            "nullRatio": round(1 - sample_not_null / len(reservoir), 6) if reservoir else 0,
            "logDataType": log_data_type,
            "entropy": round(entropy, 6),
            "normEntropy": round(norm_entropy, 6),
            "numericRatio": round(numeric_ratio, 6),
            "integerRatio": round(integer_count / len(numbers), 6) if numbers else 0,
            "minNumValue": round(min(numbers), 12) if numbers else None,
            "maxNumValue": round(max(numbers), 12) if numbers else None,
            "avgTextLength": round(sum(map(len, values)) / sample_not_null, 6) if sample_not_null else None,
            "maxTextLength": max(map(len, values)) if sample_not_null else None,
            "typeCode": label["typeCode"],
            "typeGroupCode": TYPE_GROUPS[label["typeCode"]],
            "displayTypeValue": label["typeCode"],
            "sourceId": f"UCI-{source['datasetId']}",
            "sourceDoi": source["doi"],
            "labelReason": f"Derived column profile from {source['title']} ({source['doi']})",
        }
        hash_source = json.dumps(profile, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
        profile["profileHash"] = hashlib.sha256(hash_source.encode("utf-8")).hexdigest().upper()
        profiles.append(profile)
    return profiles


def csv_rows(data: bytes, *, encoding: str = "utf-8", delimiter: str = ",", header: bool = True) -> tuple[list[str], Iterator[list[str]]]:
    text = io.StringIO(data.decode(encoding))
    reader = csv.reader(text, delimiter=delimiter, skipinitialspace=True)
    columns = next(reader) if header else []
    return columns, reader


def xlsx_rows(data: bytes) -> tuple[list[str], Iterator[list[object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is required only to rebuild the Online Retail seed profiles") from error
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    columns = [str(value) for value in next(iterator)]
    return columns, iterator


def load_real_profiles(cache_dir: Path) -> list[dict[str, object]]:
    profiles: list[dict[str, object]] = []

    with zipfile.ZipFile(cache_dir / "adult.zip") as archive:
        _, rows = csv_rows(archive.read("adult.data"), header=False)
        profiles += stream_profile_rows(rows, ADULT_COLUMNS, ADULT_LABELS, "adult", "ADULT")

    with zipfile.ZipFile(cache_dir / "bank.zip") as outer:
        with zipfile.ZipFile(io.BytesIO(outer.read("bank.zip"))) as archive:
            columns, rows = csv_rows(archive.read("bank-full.csv"), delimiter=";")
            profiles += stream_profile_rows(rows, columns, BANK_LABELS, "bank", "BANK_FULL")

    with zipfile.ZipFile(cache_dir / "bike.zip") as archive:
        columns, rows = csv_rows(archive.read("hour.csv"))
        profiles += stream_profile_rows(rows, columns, BIKE_LABELS, "bike", "BIKE_HOUR")

    with zipfile.ZipFile(cache_dir / "wine.zip") as archive:
        for member, table_name in (("winequality-red.csv", "WINE_RED"), ("winequality-white.csv", "WINE_WHITE")):
            columns, rows = csv_rows(archive.read(member), delimiter=";")
            profiles += stream_profile_rows(rows, columns, WINE_LABELS, "wine", table_name)

    with zipfile.ZipFile(cache_dir / "sms.zip") as archive:
        _, rows = csv_rows(archive.read("SMSSpamCollection"), delimiter="\t", header=False)
        profiles += stream_profile_rows(
            rows,
            ["label", "message"],
            {
                "label": field("CAT_CHAR", "VARCHAR2", "Spam class"),
                "message": field("FREE_TEXT", "VARCHAR2", "SMS message text"),
            },
            "sms",
            "SMS_SPAM",
        )

    with zipfile.ZipFile(cache_dir / "drybean.zip") as archive:
        text = archive.read("DryBeanDataset/Dry_Bean_Dataset.arff").decode("utf-8")
        lines = text.splitlines()
        columns = [
            re.match(r"@ATTRIBUTE\s+([^\s]+)", line, re.IGNORECASE).group(1)
            for line in lines
            if re.match(r"@ATTRIBUTE\s+([^\s]+)", line, re.IGNORECASE)
        ]
        labels = {
            column: field("CAT_GENERAL", "VARCHAR2", "Bean class")
            if column.lower() == "class"
            else field("NUM_CONTINUOUS", "NUMBER", column)
            for column in columns
        }
        data_index = next(index for index, line in enumerate(lines) if line.strip().lower() == "@data") + 1
        rows = csv.reader(lines[data_index:])
        profiles += stream_profile_rows(rows, columns, labels, "drybean", "DRY_BEAN")

    with zipfile.ZipFile(cache_dir / "seoulbike.zip") as archive:
        columns, rows = csv_rows(archive.read("SeoulBikeData.csv"), encoding="latin-1")
        profiles += stream_profile_rows(rows, columns, SEOUL_LABELS, "seoulbike", "SEOUL_BIKE")

    with zipfile.ZipFile(cache_dir / "retail.zip") as archive:
        columns, rows = xlsx_rows(archive.read("Online Retail.xlsx"))
        profiles += stream_profile_rows(rows, columns, RETAIL_LABELS, "retail", "ONLINE_RETAIL")

    return profiles


def synthetic_profiles() -> list[dict[str, object]]:
    templates = {
        "NUM_IDENTIFIER": ("NUMBER", "NUM", 0.995, 0.998, 1.0, 1.0, 1, 100000, 8, 12),
        "CHAR_IDENTIFIER": ("VARCHAR2", "CHR", 0.99, 0.997, 0.0, 0.0, None, None, 12, 32),
        "NUM_CONTINUOUS": ("NUMBER", "NUM", 0.72, 0.96, 1.0, 0.25, -500, 25000, 7, 18),
        "NUM_DISCRETE": ("NUMBER", "NUM", 0.08, 0.82, 1.0, 1.0, 0, 120, 3, 5),
        "CAT_GENERAL": ("VARCHAR2", "CHR", 0.015, 0.78, 0.0, 0.0, None, None, 9, 24),
        "CAT_CHAR": ("CHAR", "CHR", 0.004, 0.72, 0.0, 0.0, None, None, 1, 2),
        "CAT_ORDINAL": ("NUMBER", "NUM", 0.01, 0.76, 1.0, 1.0, 1, 12, 2, 2),
        "CAT_NUMERIC": ("NUMBER", "NUM", 0.02, 0.8, 1.0, 1.0, 0, 30, 2, 3),
        "FREE_TEXT": ("VARCHAR2", "CHR", 0.94, 0.99, 0.02, 0.1, None, None, 120, 800),
        "OTHER": ("DATE", "CHR", 0.85, 0.97, 0.0, 0.0, None, None, 19, 30),
    }
    profiles: list[dict[str, object]] = []
    for type_index, (type_code, values) in enumerate(templates.items(), start=1):
        data_type, log_type, ratio, norm_entropy, numeric_ratio, integer_ratio, min_value, max_value, avg_len, max_len = values
        for variant in range(1, 7):
            total = 1000 + variant * 137
            non_null = total - (variant % 3) * 7
            distinct_ratio = min(1.0, max(0.000001, ratio * (0.9 + variant * 0.025)))
            num_distinct = max(1, round(non_null * distinct_ratio))
            entropy_value = max(0.0, math.log(max(num_distinct, 1)) * min(1.0, norm_entropy * (0.94 + variant * 0.01)))
            profile = {
                "runSourceType": "DATA_WORK",
                "runId": 0,
                "owner": "INIT$SAMPLE",
                "tableName": f"BOUNDARY_COLTYPE_{variant:02d}",
                "columnName": f"COL_{type_code}",
                "featureVersion": "V2",
                "columnDesc": f"Balanced boundary sample: {type_code} variant {variant}",
                "columnId": type_index,
                "dataType": data_type,
                "totalRows": total,
                "nonNullRows": non_null,
                "sampleRows": total,
                "sampleNotNullRows": non_null,
                "numDistinct": num_distinct,
                "sampleDistinct": num_distinct,
                "distinctRatio": round(distinct_ratio, 6),
                "nullRatio": round(1 - non_null / total, 6),
                "logDataType": log_type,
                "entropy": round(entropy_value, 6),
                "normEntropy": round(min(1.0, norm_entropy * (0.94 + variant * 0.01)), 6),
                "numericRatio": numeric_ratio,
                "integerRatio": integer_ratio,
                "minNumValue": min_value,
                "maxNumValue": max_value,
                "avgTextLength": avg_len,
                "maxTextLength": max_len,
                "typeCode": type_code,
                "typeGroupCode": TYPE_GROUPS[type_code],
                "displayTypeValue": type_code,
                "sourceId": "INIT-BOUNDARY-V2",
                "sourceDoi": None,
                "labelReason": "Balanced synthetic boundary profile for cold-start model training",
            }
            hash_source = json.dumps(profile, sort_keys=True, separators=(",", ":"))
            profile["profileHash"] = hashlib.sha256(hash_source.encode("utf-8")).hexdigest().upper()
            profiles.append(profile)
    return profiles


def download_sources(cache_dir: Path) -> None:
    cache_dir.mkdir(parents=True, exist_ok=True)
    for key, source in SOURCES.items():
        target = cache_dir / f"{key}.zip"
        if target.exists() and target.stat().st_size:
            continue
        print(f"Downloading {source['title']} -> {target}")
        urllib.request.urlretrieve(source["url"], target)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cache-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--download", action="store_true")
    args = parser.parse_args()
    if args.download:
        download_sources(args.cache_dir)

    profiles = synthetic_profiles() + load_real_profiles(args.cache_dir)
    profiles.sort(key=lambda item: (str(item["tableName"]), int(item["columnId"]), str(item["columnName"])))
    counts = Counter(str(item["typeCode"]) for item in profiles)
    payload = {
        "schemaVersion": 1,
        "featureVersion": "V2",
        "sampleLimitPerDataset": SAMPLE_LIMIT,
        "description": "Cold-start column-type profiles derived from CC BY 4.0 UCI datasets plus balanced boundary profiles. No raw source records are included.",
        "sources": list(SOURCES.values()),
        "typeCounts": dict(sorted(counts.items())),
        "profiles": profiles,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(profiles)} profiles to {args.output}")
    print(json.dumps(payload["typeCounts"], ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
