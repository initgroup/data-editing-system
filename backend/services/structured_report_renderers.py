from __future__ import annotations

import html
import io
import json
import logging
import os
import re
import shutil
import subprocess
import threading
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.services.report_fonts import REPORT_FONT_FAMILY, embedded_korean_font_css
from backend.services.report_i18n import REPORT_UI_TEXT, normalize_report_language


_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
_ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_PDF_SEMAPHORE = threading.BoundedSemaphore(2)
_BATCH_PDF_SEMAPHORE = threading.BoundedSemaphore(1)
_PDF_MAX_HTML_BYTES = 15 * 1024 * 1024
_BATCH_PDF_MAX_HTML_BYTES = 40 * 1024 * 1024
_PDF_TIMEOUT_SECONDS = 45
_BATCH_PDF_WATCHDOG_MS = 120_000
_BATCH_PDF_CONTENT_TIMEOUT_MS = 60_000
_BATCH_PDF_TIMEOUT_SECONDS = 135
_ROOT_DIR = Path(__file__).resolve().parents[2]
_PDF_SCRIPT = _ROOT_DIR / "scripts" / "render_report_pdf.mjs"
_LOCAL_PLAYWRIGHT_BROWSERS = _ROOT_DIR / "node_modules" / "playwright-core" / ".local-browsers"
_KST = timezone(timedelta(hours=9), name="Asia/Seoul")
logger = logging.getLogger(__name__)


def _report_language(document: dict[str, Any]) -> str:
    return normalize_report_language(document.get("language"))


def _report_labels(document: dict[str, Any]) -> dict[str, str]:
    language = _report_language(document)
    labels = dict(REPORT_UI_TEXT[language])
    supplied = document.get("labels")
    if isinstance(supplied, dict):
        labels.update({str(key): str(value) for key, value in supplied.items() if value is not None})
    return labels


def _value_text(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, Decimal):
        if not value.is_finite():
            return str(value)
        text = format(value, ",f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    if isinstance(value, float):
        if abs(value) <= 1 and value != 0:
            return f"{value:.4f}".rstrip("0").rstrip(".")
        return f"{value:,.4f}".rstrip("0").rstrip(".")
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            if value.tzinfo is not None and value.utcoffset() is not None:
                value = value.astimezone(_KST)
            return value.isoformat(sep=" ", timespec="seconds")
        return value.isoformat()
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _format_kpi_value(kpi: dict[str, Any]) -> str:
    value = kpi.get("value")
    if value is None:
        return "-"
    if kpi.get("unit") == "RATE":
        return f"{float(value) * 100:,.2f}%"
    return _value_text(value)


def _escape(value: Any) -> str:
    return html.escape(_value_text(value), quote=True)


def _context_name(document: dict[str, Any]) -> tuple[str, str]:
    context = document.get("context") or {}
    project = context.get("project") or {}
    scenario = context.get("scenario") or {}
    project_name = project.get("PROJECT_NAME") or project.get("PROJECT_CODE") or "-"
    scenario_name = scenario.get("SCENARIO_NAME") or scenario.get("SCENARIO_CODE") or _report_labels(document)["allScenarios"]
    return str(project_name), str(scenario_name)


def render_report_html(document: dict[str, Any], *, embed_fonts: bool = False) -> bytes:
    report = document.get("report") or {}
    provider = document.get("provider") or {}
    context = document.get("context") or {}
    selection = context.get("selection") or {}
    project_name, scenario_name = _context_name(document)
    availability = document.get("availability") or {}
    kpis = document.get("kpis") or []
    sections = document.get("sections") or []
    definitions = document.get("definitions") or []
    language = _report_language(document)
    labels = _report_labels(document)
    font_css = embedded_korean_font_css() if embed_fonts else ""
    max_table_columns = max(
        (len(section.get("columns") or []) for section in sections if section.get("type") != "text"),
        default=0,
    )
    page_size = "A4 landscape" if max_table_columns > 10 else "A4"
    page_margin = "10mm 7mm 12mm" if max_table_columns > 10 else "14mm 12mm 16mm"

    kpi_html = []
    for item in kpis:
        ratio_hint = ""
        if item.get("numerator") is not None or item.get("denominator") is not None:
            ratio_hint = (
                f'<div class="kpi-ratio">{_escape(labels["numerator"])} {_escape(item.get("numerator"))} / '
                f'{_escape(labels["denominator"])} {_escape(item.get("denominator"))}</div>'
            )
        kpi_html.append(
            "<article class=\"kpi\">"
            f"<div class=\"kpi-label\">{_escape(item.get('label'))}</div>"
            f"<div class=\"kpi-value\">{_escape(_format_kpi_value(item))}</div>"
            f"{ratio_hint}</article>"
        )

    section_html = []
    for section in sections:
        title = _escape(section.get("title"))
        description = section.get("description")
        note = section.get("note")
        if section.get("type") == "text":
            paragraphs = "".join(f"<p>{_escape(item)}</p>" for item in section.get("paragraphs") or [])
            body = f'<div class="text-section">{paragraphs}</div>'
        else:
            columns = section.get("columns") or []
            rows = section.get("rows") or []
            column_count = len(columns)
            table_wrap_class = "table-wrap table-wide" if column_count > 10 else "table-wrap"
            header = "".join(
                f"<th scope=\"col\">{_escape(column.get('label') or column.get('key'))}</th>"
                for column in columns
            )
            table_rows = []
            for row in rows:
                cells = "".join(
                    f'<td data-label="{_escape(column.get("label") or column.get("key"))}">'
                    f"{_escape(row.get(column.get('key')))}</td>"
                    for column in columns
                )
                table_rows.append(f"<tr>{cells}</tr>")
            empty = "" if rows else f'<div class="empty">{_escape(availability.get("reason") or labels["noData"])}</div>'
            table = (
                f'<div class="{table_wrap_class}" data-column-count="{column_count}"><table><thead><tr>{header}</tr></thead>'
                f'<tbody>{"".join(table_rows)}</tbody></table></div>'
                if rows
                else ""
            )
            body = f"{table}{empty}"
        description_html = f'<p class="section-description">{_escape(description)}</p>' if description else ""
        note_html = f'<p class="section-note">{_escape(note)}</p>' if note else ""
        section_html.append(
            f'<section class="report-section"><h2>{title}</h2>{description_html}{body}{note_html}</section>'
        )

    definition_rows = "".join(
        f"<tr><th>{_escape(item.get('term'))}</th><td>{_escape(item.get('definition'))}</td></tr>"
        for item in definitions
    )

    result = f"""<!doctype html>
<html lang="{_escape(language)}">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta http-equiv="Content-Security-Policy" content="default-src 'none'; style-src 'unsafe-inline'; font-src data:; img-src data:">
  <title>{_escape(provider.get('name'))} {_escape(report.get('code'))} {_escape(report.get('title'))}</title>
  <style>
    {font_css}
    @page {{ size: {page_size}; margin: {page_margin}; }}
    :root {{ color-scheme: light; --ink:#172033; --muted:#64748b; --line:#dbe3ee; --brand:#0f766e; --soft:#f4f8fb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; color:var(--ink); background:#eef3f7; font-family:"{REPORT_FONT_FAMILY}","Noto Sans KR","Malgun Gothic","Apple SD Gothic Neo",sans-serif; line-height:1.55; }}
    .page {{ width:min(1180px, calc(100% - 32px)); margin:24px auto; padding:42px; background:#fff; border:1px solid var(--line); border-radius:18px; box-shadow:0 18px 50px rgba(15,23,42,.08); }}
    .brand {{ display:flex; justify-content:space-between; gap:18px; align-items:flex-start; padding-bottom:22px; border-bottom:3px solid var(--brand); }}
    .brand-name {{ color:var(--brand); font-size:14px; font-weight:900; letter-spacing:.12em; }}
    h1 {{ margin:6px 0 0; font-size:28px; line-height:1.3; }}
    .report-code {{ padding:8px 12px; border-radius:999px; background:#e6fffb; color:#115e59; font-weight:800; white-space:nowrap; }}
    .provider-statement {{ margin:14px 0 0; color:var(--muted); font-size:13px; }}
    .meta {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin:22px 0; }}
    .meta-item {{ min-width:0; padding:12px 14px; border-radius:12px; background:var(--soft); }}
    .meta-label {{ display:block; color:var(--muted); font-size:11px; font-weight:700; }}
    .meta-value {{ display:block; margin-top:4px; overflow-wrap:anywhere; font-size:13px; font-weight:700; }}
    .availability {{ margin:0 0 20px; padding:10px 13px; border:1px solid #bae6fd; border-radius:10px; background:#f0f9ff; color:#075985; font-size:12px; }}
    .kpis {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; margin:20px 0 28px; }}
    .kpi {{ min-width:0; padding:17px; border:1px solid var(--line); border-radius:14px; background:linear-gradient(145deg,#fff,#f7fafc); }}
    .kpi-label {{ color:var(--muted); font-size:12px; font-weight:700; }}
    .kpi-value {{ margin-top:7px; font-size:24px; font-weight:900; color:#0f4c5c; overflow-wrap:anywhere; }}
    .kpi-ratio {{ margin-top:4px; color:var(--muted); font-size:10px; }}
    .report-section {{ margin:28px 0; break-inside:auto; }}
    h2 {{ margin:0 0 12px; padding-left:10px; border-left:4px solid var(--brand); font-size:18px; }}
    .section-description,.section-note {{ color:var(--muted); font-size:12px; }}
    .section-note {{ margin-top:9px; padding:10px 12px; border-radius:9px; background:#fff7ed; color:#9a3412; }}
    .text-section p {{ margin:7px 0; }}
    .table-wrap {{ width:100%; overflow:visible; border:1px solid var(--line); border-radius:10px; }}
    table {{ width:100%; border-collapse:collapse; table-layout:fixed; font-size:11px; }}
    thead {{ display:table-header-group; }}
    tr {{ break-inside:avoid; }}
    th,td {{ min-width:0; padding:8px 9px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; overflow-wrap:anywhere; word-break:break-word; white-space:normal; }}
    thead th {{ background:#eef5f7; color:#334155; font-weight:800; }}
    .table-wide table {{ font-size:9px; }}
    .table-wide th,.table-wide td {{ padding:6px 5px; }}
    tbody tr:nth-child(even) {{ background:#fafcfd; }}
    .empty {{ padding:24px; text-align:center; color:var(--muted); background:var(--soft); border-radius:10px; }}
    .definitions th {{ width:150px; }}
    footer {{ margin-top:34px; padding-top:14px; border-top:1px solid var(--line); color:var(--muted); font-size:10px; text-align:center; }}
    @media (max-width:760px) {{
      .page {{ width:100%; margin:0; padding:22px 16px; border:0; border-radius:0; box-shadow:none; }}
      .brand {{ display:block; }} .report-code {{ display:inline-block; margin-top:12px; }}
      h1 {{ font-size:23px; }} .meta {{ grid-template-columns:1fr 1fr; }} .kpis {{ grid-template-columns:1fr 1fr; }}
      .table-wrap[data-column-count] {{ border:0; }}
      .table-wrap[data-column-count] table,.table-wrap[data-column-count] tbody,.table-wrap[data-column-count] tr,.table-wrap[data-column-count] td {{ display:block; width:100%; }}
      .table-wrap[data-column-count] table {{ table-layout:auto; }}
      .table-wrap[data-column-count] thead {{ clip:rect(0 0 0 0); clip-path:inset(50%); height:1px; overflow:hidden; position:absolute; white-space:nowrap; width:1px; }}
      .table-wrap[data-column-count] tbody {{ display:grid; gap:10px; }}
      .table-wrap[data-column-count] tr {{ border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
      .table-wrap[data-column-count] td {{ display:grid; grid-template-columns:minmax(96px,34%) minmax(0,1fr); gap:12px; padding:9px 10px; border:0; border-bottom:1px solid var(--line); font-size:11px; }}
      .table-wrap[data-column-count] td:last-child {{ border-bottom:0; }}
      .table-wrap[data-column-count] td::before {{ content:attr(data-label); color:var(--muted); font-size:10px; font-weight:800; overflow-wrap:anywhere; }}
    }}
    @media (max-width:420px) {{
      .meta,.kpis {{ grid-template-columns:1fr; }} .kpi-value {{ font-size:21px; }}
      .table-wrap[data-column-count] td {{ grid-template-columns:1fr; gap:4px; }}
    }}
    @media print {{
      body {{ background:#fff; }} .page {{ width:100%; margin:0; padding:0; border:0; border-radius:0; box-shadow:none; }}
      .table-wrap {{ max-width:100%; overflow:visible; }}
      .table-wrap table {{ width:100% !important; max-width:100%; table-layout:fixed; font-size:8px; }}
      .table-wrap th,.table-wrap td {{ padding:4px 3px; line-height:1.35; overflow-wrap:anywhere; word-break:break-word; white-space:normal !important; }}
      .table-wrap.table-wide table {{ font-size:7px; }}
      .table-wrap.table-wide th,.table-wrap.table-wide td {{ padding:3px 2px; line-height:1.25; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <header class="brand">
      <div>
        <div class="brand-name">{_escape(provider.get('name') or 'IN-DEPS')}</div>
        <h1>{_escape(report.get('title'))}</h1>
        <p class="provider-statement">{_escape(provider.get('statement'))}</p>
      </div>
      <div class="report-code">{_escape(report.get('code'))}</div>
    </header>
    <div class="meta">
      <div class="meta-item"><span class="meta-label">{_escape(labels['project'])}</span><span class="meta-value">{_escape(project_name)}</span></div>
      <div class="meta-item"><span class="meta-label">{_escape(labels['scenario'])}</span><span class="meta-value">{_escape(scenario_name)}</span></div>
      <div class="meta-item"><span class="meta-label">{_escape(labels['flowRunId'])}</span><span class="meta-value">{_escape(selection.get('flowRunId'))}</span></div>
      <div class="meta-item"><span class="meta-label">{_escape(labels['editingSessionId'])}</span><span class="meta-value">{_escape(selection.get('editSessionId'))}</span></div>
      <div class="meta-item"><span class="meta-label">{_escape(labels['definitionVersion'])}</span><span class="meta-value">{_escape(report.get('definitionVersion'))}</span></div>
      <div class="meta-item"><span class="meta-label">{_escape(labels['generatedAt'])}</span><span class="meta-value">{_escape(context.get('generatedAt'))}</span></div>
    </div>
    <div class="availability">{_escape(labels['status'])}: {_escape(availability.get('label') or availability.get('status'))}{' · ' + _escape(availability.get('reason')) if availability.get('reason') else ''}</div>
    <section class="kpis">{"".join(kpi_html)}</section>
    {"".join(section_html)}
    <section class="report-section definitions">
      <h2>{_escape(labels['definitions'])}</h2>
      <div class="table-wrap"><table><tbody>{definition_rows}</tbody></table></div>
    </section>
    <footer>{_escape(provider.get('statement'))} · {_escape(report.get('definitionVersion'))}</footer>
  </main>
</body>
</html>"""
    return result.encode("utf-8")


def _bundle_status_class(value: Any) -> str:
    status = str(value or "NO_DATA").upper()
    return {
        "AVAILABLE": "status-available",
        "PARTIAL": "status-partial",
        "NO_DATA": "status-empty",
        "NOT_APPLICABLE": "status-empty",
        "ERROR": "status-error",
    }.get(status, "status-empty")


def _render_bundle_kpis(kpis: list[dict[str, Any]], labels: dict[str, str]) -> str:
    if not kpis:
        return ""
    items = []
    for item in kpis:
        ratio_hint = ""
        if item.get("numerator") is not None or item.get("denominator") is not None:
            ratio_hint = (
                f'<div class="kpi-ratio">{_escape(labels["numerator"])} {_escape(item.get("numerator"))} / '
                f'{_escape(labels["denominator"])} {_escape(item.get("denominator"))}</div>'
            )
        items.append(
            '<article class="kpi">'
            f'<div class="kpi-label">{_escape(item.get("label"))}</div>'
            f'<div class="kpi-value">{_escape(_format_kpi_value(item))}</div>'
            f"{ratio_hint}</article>"
        )
    return f'<section class="kpis">{"".join(items)}</section>'


def _render_bundle_sections(document: dict[str, Any]) -> str:
    availability = document.get("availability") or {}
    labels = _report_labels(document)
    rendered = []
    for section in document.get("sections") or []:
        title = _escape(section.get("title"))
        if section.get("type") == "text":
            paragraphs = "".join(f"<p>{_escape(item)}</p>" for item in section.get("paragraphs") or [])
            body = f'<div class="text-section">{paragraphs}</div>'
        else:
            columns = section.get("columns") or []
            rows = section.get("rows") or []
            column_count = len(columns)
            table_class = "table-wrap table-wide" if column_count > 10 else "table-wrap"
            header = "".join(
                f'<th scope="col">{_escape(column.get("label") or column.get("key"))}</th>'
                for column in columns
            )
            table_rows = []
            for row in rows:
                cells = "".join(
                    f'<td data-label="{_escape(column.get("label") or column.get("key"))}">'
                    f'{_escape(row.get(column.get("key")))}</td>'
                    for column in columns
                )
                table_rows.append(f'<tr>{cells}</tr>')
            body = (
                f'<div class="{table_class}" data-column-count="{column_count}"><table>'
                f'<thead><tr>{header}</tr></thead><tbody>{"".join(table_rows)}</tbody></table></div>'
                if rows
                else f'<div class="empty">{_escape(availability.get("reason") or labels["noData"])}</div>'
            )
        description = (
            f'<p class="section-description">{_escape(section.get("description"))}</p>'
            if section.get("description")
            else ""
        )
        note = (
            f'<p class="section-note">{_escape(section.get("note"))}</p>'
            if section.get("note")
            else ""
        )
        rendered.append(
            f'<section class="report-section"><h3>{title}</h3>{description}{body}{note}</section>'
        )
    return "".join(rendered)


def render_report_bundle_html(document: dict[str, Any], *, embed_fonts: bool = False) -> bytes:
    provider = document.get("provider") or {}
    bundle = document.get("bundle") or {}
    context = document.get("context") or {}
    selection = context.get("selection") or {}
    summary = document.get("summary") or {}
    reports = document.get("reports") or []
    project_name, scenario_name = _context_name(document)
    language = _report_language(document)
    labels = _report_labels(document)
    font_css = embedded_korean_font_css() if embed_fonts else ""

    summary_items = [
        (labels["total"], summary.get("totalCount")),
        (labels["available"], summary.get("availableCount")),
        (labels["partial"], summary.get("partialCount")),
        (labels["noData"], summary.get("noDataCount")),
        (labels["notApplicable"], summary.get("notApplicableCount")),
        (labels["error"], summary.get("errorCount")),
    ]
    summary_html = "".join(
        f'<article><span>{_escape(label)}</span><strong>{_escape(value)}</strong></article>'
        for label, value in summary_items
    )

    toc_items = []
    chapters = []
    for index, report_document in enumerate(reports, start=1):
        report = report_document.get("report") or {}
        availability = report_document.get("availability") or {}
        status = str(availability.get("status") or "NO_DATA").upper()
        anchor = f"report-{index:02d}"
        toc_items.append(
            f'<li><a href="#{anchor}"><span class="toc-code">{_escape(report.get("code"))}</span>'
            f'<span class="toc-title">{_escape(report.get("title"))}</span>'
            f'<span class="status {_bundle_status_class(status)}">{_escape(availability.get("label") or status)}</span></a></li>'
        )
        definitions = report_document.get("definitions") or []
        definition_rows = "".join(
            f'<tr><th scope="row">{_escape(item.get("term"))}</th><td>{_escape(item.get("definition"))}</td></tr>'
            for item in definitions
        )
        reason = (
            f' · {_escape(availability.get("reason"))}'
            if availability.get("reason")
            else ""
        )
        chapters.append(
            f'<article class="report-chapter" id="{anchor}">'
            '<header class="chapter-header">'
            f'<div><span class="chapter-number">{index:02d}</span>'
            f'<span class="chapter-code">{_escape(report.get("code"))}</span></div>'
            f'<h2>{_escape(report.get("title"))}</h2>'
            f'<p>{_escape(report.get("description"))}</p>'
            f'<div class="availability"><span class="status {_bundle_status_class(status)}">{_escape(availability.get("label") or status)}</span>{reason}</div>'
            '</header>'
            f'{_render_bundle_kpis(report_document.get("kpis") or [], labels)}'
            f'{_render_bundle_sections(report_document)}'
            f'<section class="report-section definitions"><h3>{_escape(labels["definitions"])}</h3>'
            f'<div class="table-wrap"><table><tbody>{definition_rows}</tbody></table></div></section>'
            f'<div class="chapter-footer"><a href="#report-index">{_escape(labels["backToToc"])}</a></div>'
            '</article>'
        )

    result = f"""<!doctype html>
<html lang="{_escape(language)}">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta http-equiv="Content-Security-Policy" content="default-src 'none'; style-src 'unsafe-inline'; font-src data:; img-src data:">
  <title>{_escape(provider.get('name'))} {_escape(bundle.get('title'))}</title>
  <style>
    {font_css}
    @page {{ size:A4 landscape; margin:10mm 8mm 12mm; }}
    :root {{ color-scheme:light; --ink:#172033; --muted:#64748b; --line:#dbe3ee; --brand:#174a87; --soft:#f3f7fb; }}
    * {{ box-sizing:border-box; min-width:0; }}
    html,body {{ max-width:100%; overflow-x:hidden; }}
    body {{ margin:0; color:var(--ink); background:#eef3f7; font-family:"{REPORT_FONT_FAMILY}","Noto Sans KR","Malgun Gothic","Apple SD Gothic Neo",sans-serif; line-height:1.5; }}
    .page {{ width:min(1240px,calc(100% - 32px)); margin:24px auto; padding:40px; background:#fff; border:1px solid var(--line); border-radius:18px; box-shadow:0 18px 50px rgba(15,23,42,.08); }}
    .bundle-header {{ padding-bottom:24px; border-bottom:4px solid var(--brand); }}
    .brand-name {{ color:var(--brand); font-size:13px; font-weight:900; letter-spacing:.12em; }}
    h1 {{ margin:7px 0; font-size:30px; }}
    .provider-statement,.bundle-description {{ color:var(--muted); font-size:13px; }}
    .meta {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin:20px 0; }}
    .meta div {{ padding:12px; border-radius:10px; background:var(--soft); }}
    .meta span {{ display:block; color:var(--muted); font-size:10px; font-weight:700; }}
    .meta strong {{ display:block; margin-top:4px; overflow-wrap:anywhere; font-size:12px; }}
    .summary-grid {{ display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:10px; margin:20px 0 28px; }}
    .summary-grid article {{ padding:13px; border:1px solid var(--line); border-radius:12px; text-align:center; }}
    .summary-grid span {{ display:block; color:var(--muted); font-size:10px; }}
    .summary-grid strong {{ display:block; margin-top:4px; color:var(--brand); font-size:22px; }}
    .toc {{ margin:30px 0; padding:22px; background:var(--soft); border-radius:14px; }}
    .toc h2 {{ margin:0 0 13px; font-size:20px; }}
    .toc ol {{ display:grid; grid-template-columns:1fr 1fr; gap:7px 20px; margin:0; padding:0; list-style:none; }}
    .toc a {{ display:grid; grid-template-columns:42px minmax(0,1fr) auto; gap:8px; align-items:center; padding:7px; border-bottom:1px solid var(--line); color:inherit; text-decoration:none; }}
    .toc-code,.chapter-code {{ color:var(--brand); font-size:11px; font-weight:900; }}
    .toc-title {{ overflow-wrap:anywhere; font-size:11px; font-weight:700; }}
    .status {{ display:inline-block; padding:3px 7px; border-radius:999px; font-size:9px; font-weight:900; white-space:nowrap; }}
    .status-available {{ color:#166534; background:#dcfce7; }} .status-partial {{ color:#92400e; background:#fef3c7; }}
    .status-empty {{ color:#475569; background:#e2e8f0; }} .status-error {{ color:#991b1b; background:#fee2e2; }}
    .report-chapter {{ padding-top:24px; scroll-margin-top:12px; break-before:page; page-break-before:always; }}
    .chapter-header {{ padding-bottom:15px; border-bottom:2px solid var(--brand); }}
    .chapter-number {{ display:inline-grid; place-items:center; width:30px; height:30px; margin-right:9px; border-radius:8px; color:#fff; background:var(--brand); font-size:12px; font-weight:900; }}
    .chapter-header h2 {{ margin:10px 0 4px; font-size:23px; }} .chapter-header p {{ margin:0 0 8px; color:var(--muted); font-size:12px; }}
    .availability {{ overflow-wrap:anywhere; color:var(--muted); font-size:10px; }}
    .kpis {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin:18px 0 24px; }}
    .kpi {{ padding:13px; border:1px solid var(--line); border-radius:12px; background:#fff; }}
    .kpi-label {{ color:var(--muted); font-size:10px; font-weight:700; }} .kpi-value {{ margin-top:5px; color:var(--brand); font-size:20px; font-weight:900; overflow-wrap:anywhere; }}
    .kpi-ratio {{ color:var(--muted); font-size:9px; }}
    .report-section {{ margin:24px 0; }} .report-section h3 {{ margin:0 0 10px; padding-left:9px; border-left:4px solid var(--brand); font-size:16px; }}
    .section-description,.section-note {{ color:var(--muted); font-size:10px; }}
    .section-note {{ padding:8px 10px; border-radius:8px; color:#9a3412; background:#fff7ed; }}
    .text-section p {{ margin:6px 0; font-size:11px; }}
    .table-wrap {{ width:100%; max-width:100%; overflow:visible; border:1px solid var(--line); border-radius:9px; }}
    table {{ width:100%; max-width:100%; border-collapse:collapse; table-layout:fixed; font-size:9px; }}
    thead {{ display:table-header-group; }} tr {{ break-inside:avoid; }}
    th,td {{ min-width:0; padding:6px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; overflow-wrap:anywhere; word-break:break-word; white-space:normal; }}
    thead th {{ color:#334155; background:#eaf1f8; font-weight:900; }}
    .table-wide table {{ font-size:7px; }} .table-wide th,.table-wide td {{ padding:4px 3px; }}
    tbody tr:nth-child(even) {{ background:#fafcfd; }} .definitions th {{ width:18%; }}
    .empty {{ padding:18px; border-radius:9px; color:var(--muted); background:var(--soft); text-align:center; }}
    .chapter-footer {{ margin-top:18px; text-align:right; }} .chapter-footer a {{ color:var(--brand); font-size:10px; }}
    .bundle-footer {{ margin-top:30px; padding-top:14px; border-top:1px solid var(--line); color:var(--muted); font-size:10px; text-align:center; }}
    @media (max-width:760px) {{
      .page {{ width:100%; margin:0; padding:20px 14px; border:0; border-radius:0; box-shadow:none; }} h1 {{ font-size:24px; }}
      .meta {{ grid-template-columns:1fr 1fr; }} .summary-grid {{ grid-template-columns:repeat(3,1fr); }} .toc ol {{ grid-template-columns:1fr; }}
      .kpis {{ grid-template-columns:1fr 1fr; }}
      .table-wrap[data-column-count] {{ border:0; }}
      .table-wrap[data-column-count] table,.table-wrap[data-column-count] tbody,.table-wrap[data-column-count] tr,.table-wrap[data-column-count] td {{ display:block; width:100%; }}
      .table-wrap[data-column-count] table {{ table-layout:auto; }}
      .table-wrap[data-column-count] thead {{ clip:rect(0 0 0 0); clip-path:inset(50%); height:1px; overflow:hidden; position:absolute; white-space:nowrap; width:1px; }}
      .table-wrap[data-column-count] tbody {{ display:grid; gap:9px; }}
      .table-wrap[data-column-count] tr {{ overflow:hidden; border:1px solid var(--line); border-radius:9px; }}
      .table-wrap[data-column-count] td {{ display:grid; grid-template-columns:minmax(96px,34%) minmax(0,1fr); gap:10px; padding:8px 9px; border:0; border-bottom:1px solid var(--line); font-size:10px; }}
      .table-wrap[data-column-count] td:last-child {{ border-bottom:0; }}
      .table-wrap[data-column-count] td::before {{ content:attr(data-label); color:var(--muted); font-size:9px; font-weight:900; overflow-wrap:anywhere; }}
    }}
    @media (max-width:420px) {{ .meta,.kpis {{ grid-template-columns:1fr; }} .summary-grid {{ grid-template-columns:1fr 1fr; }} .toc a {{ grid-template-columns:38px minmax(0,1fr); }} .toc .status {{ grid-column:2; justify-self:start; }} .table-wrap[data-column-count] td {{ grid-template-columns:1fr; gap:3px; }} }}
    @media print {{
      body {{ background:#fff; }} .page {{ width:100%; margin:0; padding:0; border:0; border-radius:0; box-shadow:none; }}
      .report-chapter:first-of-type {{ break-before:auto; page-break-before:auto; }} .chapter-footer {{ display:none; }}
      table {{ width:100% !important; max-width:100%; table-layout:fixed; }} th,td {{ overflow-wrap:anywhere; word-break:break-word; white-space:normal !important; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <header class="bundle-header">
      <div class="brand-name">{_escape(provider.get('name') or 'IN-DEPS')}</div>
      <h1>{_escape(bundle.get('title') or labels['basicReportBundle'])}</h1>
      <p class="bundle-description">{_escape(bundle.get('description'))}</p>
      <p class="provider-statement">{_escape(provider.get('statement'))}</p>
    </header>
    <section class="meta">
      <div><span>{_escape(labels['project'])}</span><strong>{_escape(project_name)}</strong></div>
      <div><span>{_escape(labels['scenario'])}</span><strong>{_escape(scenario_name)}</strong></div>
      <div><span>{_escape(labels['flowRunId'])}</span><strong>{_escape(selection.get('flowRunId'))}</strong></div>
      <div><span>{_escape(labels['editingSessionId'])}</span><strong>{_escape(selection.get('editSessionId'))}</strong></div>
      <div><span>{_escape(labels['definitionVersion'])}</span><strong>{_escape(bundle.get('definitionVersion'))}</strong></div>
      <div><span>{_escape(labels['generatedAt'])}</span><strong>{_escape(context.get('generatedAt'))}</strong></div>
      <div><span>{_escape(labels['generationStatus'])}</span><strong>{_escape(summary.get('generationStatus'))}</strong></div>
    </section>
    <section class="summary-grid">{summary_html}</section>
    <nav class="toc" id="report-index" aria-label="{_escape(labels['toc'])}"><h2>{_escape(labels['toc'])}</h2><ol>{''.join(toc_items)}</ol></nav>
    {''.join(chapters)}
    <footer class="bundle-footer">{_escape(provider.get('statement'))} · {_escape(bundle.get('definitionVersion'))}</footer>
  </main>
</body>
</html>"""
    return result.encode("utf-8")


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            value = value.astimezone(_KST).replace(tzinfo=None)
        return value
    if value is None or isinstance(value, (bool, int, float, date)):
        return value
    if isinstance(value, Decimal):
        if value.is_finite() and len(value.as_tuple().digits) <= 15:
            return float(value)
        value = format(value, "f") if value.is_finite() else str(value)
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    text = _ILLEGAL_XML_CHARACTERS.sub("", str(value))
    if text.startswith(_FORMULA_PREFIXES):
        text = "'" + text
    return text[:32767]


def _style_header(row) -> None:
    fill = PatternFill("solid", fgColor="0F766E")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize_sheet(sheet, *, max_width: int = 48) -> None:
    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        width = 10
        for cell in sheet[letter][: min(sheet.max_row, 300)]:
            width = max(width, min(max_width, len(_value_text(cell.value)) + 2))
        sheet.column_dimensions[letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def render_report_xlsx(document: dict[str, Any]) -> bytes:
    workbook = Workbook()
    labels = _report_labels(document)
    summary_sheet = workbook.active
    summary_sheet.title = labels["summary"]
    detail_sheet = workbook.create_sheet(labels["detail"])
    definition_sheet = workbook.create_sheet(labels["calculationBasis"])
    metadata_sheet = workbook.create_sheet(labels["metadata"])
    report = document.get("report") or {}
    provider = document.get("provider") or {}
    context = document.get("context") or {}
    selection = context.get("selection") or {}
    project_name, scenario_name = _context_name(document)

    summary_sheet.append([labels["basicReport"], _xlsx_value(report.get("title"))])
    summary_sheet.append([labels["reportCode"], _xlsx_value(report.get("code"))])
    summary_sheet.append([labels["providerStatement"], _xlsx_value(provider.get("statement"))])
    summary_sheet.append([labels["project"], _xlsx_value(project_name)])
    summary_sheet.append([labels["scenario"], _xlsx_value(scenario_name)])
    summary_sheet.append([labels["generatedAt"], _xlsx_value(context.get("generatedAt"))])
    summary_sheet.append([])
    summary_sheet.append(["KPI Code", "KPI", labels["value"], labels["unit"], labels["numerator"], labels["denominator"]])
    _style_header(summary_sheet[summary_sheet.max_row])
    for kpi in document.get("kpis") or []:
        summary_sheet.append(
            [
                _xlsx_value(kpi.get("code")),
                _xlsx_value(kpi.get("label")),
                _xlsx_value(kpi.get("value")),
                _xlsx_value(kpi.get("unit")),
                _xlsx_value(kpi.get("numerator")),
                _xlsx_value(kpi.get("denominator")),
            ]
        )

    for section in document.get("sections") or []:
        detail_sheet.append([_xlsx_value(section.get("title"))])
        title_cell = detail_sheet.cell(detail_sheet.max_row, 1)
        title_cell.font = Font(bold=True, size=13, color="0F766E")
        if section.get("description"):
            detail_sheet.append([_xlsx_value(section.get("description"))])
        if section.get("type") == "text":
            for paragraph in section.get("paragraphs") or []:
                detail_sheet.append([_xlsx_value(paragraph)])
        else:
            columns = section.get("columns") or []
            if columns:
                detail_sheet.append([_xlsx_value(column.get("label") or column.get("key")) for column in columns])
                _style_header(detail_sheet[detail_sheet.max_row])
                for row in section.get("rows") or []:
                    detail_sheet.append([_xlsx_value(row.get(column.get("key"))) for column in columns])
            else:
                detail_sheet.append([labels["noData"]])
        if section.get("note"):
            detail_sheet.append([labels["notice"], _xlsx_value(section.get("note"))])
        detail_sheet.append([])

    definition_sheet.append([labels["term"], labels["definition"]])
    _style_header(definition_sheet[1])
    for item in document.get("definitions") or []:
        definition_sheet.append([_xlsx_value(item.get("term")), _xlsx_value(item.get("definition"))])

    metadata_rows = [
        (labels["providerSystem"], provider.get("name")),
        (labels["providerStatement"], provider.get("statement")),
        (labels["schemaVersion"], document.get("schemaVersion")),
        (labels["definitionVersion"], report.get("definitionVersion")),
        (labels["reportCode"], report.get("code")),
        (f'{labels["project"]} ID', selection.get("projectId")),
        (labels["project"], project_name),
        (f'{labels["scenario"]} ID', selection.get("scenarioId")),
        (labels["scenario"], scenario_name),
        (labels["flowRunId"], selection.get("flowRunId")),
        (labels["editingSessionId"], selection.get("editSessionId")),
        (labels["generatedAt"], context.get("generatedAt")),
        (labels["availabilityStatus"], (document.get("availability") or {}).get("label") or (document.get("availability") or {}).get("status")),
        (labels["availabilityReason"], (document.get("availability") or {}).get("reason")),
    ]
    metadata_sheet.append([labels["item"], labels["value"]])
    _style_header(metadata_sheet[1])
    for key, value in metadata_rows:
        metadata_sheet.append([_xlsx_value(key), _xlsx_value(value)])

    thin = Side(style="thin", color="D9E2EC")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
        _autosize_sheet(sheet)
    summary_sheet.freeze_panes = "A9"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _bundle_sheet_title(report: dict[str, Any], used_names: set[str]) -> str:
    code = str(report.get("code") or "REPORT").strip().upper()
    title = str(report.get("title") or "보고서").strip()
    base = re.sub(r"[\\/*?:\[\]]+", "_", f"{code} {title}").strip(" '")[:31] or code[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used_names:
        marker = f"_{suffix}"
        candidate = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    used_names.add(candidate.casefold())
    return candidate


def _append_bundle_report_sheet(sheet, document: dict[str, Any], provider: dict[str, Any]) -> None:
    report = document.get("report") or {}
    availability = document.get("availability") or {}
    labels = _report_labels(document)
    sheet.append([labels["basicReport"], _xlsx_value(report.get("title"))])
    sheet.append([labels["reportCode"], _xlsx_value(report.get("code"))])
    sheet.append([labels["description"], _xlsx_value(report.get("description"))])
    sheet.append([labels["status"], _xlsx_value(availability.get("label") or availability.get("status"))])
    sheet.append([labels["availabilityReason"], _xlsx_value(availability.get("reason"))])
    sheet.append([labels["dataCount"], _xlsx_value(availability.get("dataCount"))])
    sheet.append([labels["providerStatement"], _xlsx_value(provider.get("statement"))])
    sheet.append([])

    sheet.append(["KPI Code", "KPI", labels["value"], labels["unit"], labels["numerator"], labels["denominator"]])
    _style_header(sheet[sheet.max_row])
    kpis = document.get("kpis") or []
    if kpis:
        for kpi in kpis:
            sheet.append(
                [
                    _xlsx_value(kpi.get("code")),
                    _xlsx_value(kpi.get("label")),
                    _xlsx_value(kpi.get("value")),
                    _xlsx_value(kpi.get("unit")),
                    _xlsx_value(kpi.get("numerator")),
                    _xlsx_value(kpi.get("denominator")),
                ]
            )
    else:
        sheet.append([labels["noKpi"]])
    sheet.append([])

    for section in document.get("sections") or []:
        sheet.append([_xlsx_value(section.get("title"))])
        title_cell = sheet.cell(sheet.max_row, 1)
        title_cell.font = Font(bold=True, size=13, color="174A87")
        if section.get("description"):
            sheet.append([_xlsx_value(section.get("description"))])
        if section.get("type") == "text":
            paragraphs = section.get("paragraphs") or []
            if paragraphs:
                for paragraph in paragraphs:
                    sheet.append([_xlsx_value(paragraph)])
            else:
                sheet.append([labels["noContent"]])
        else:
            columns = section.get("columns") or []
            rows = section.get("rows") or []
            if columns and rows:
                sheet.append([_xlsx_value(column.get("label") or column.get("key")) for column in columns])
                _style_header(sheet[sheet.max_row])
                for row in rows:
                    sheet.append([_xlsx_value(row.get(column.get("key"))) for column in columns])
            else:
                sheet.append([labels["noData"]])
        if section.get("note"):
            sheet.append([labels["notice"], _xlsx_value(section.get("note"))])
        sheet.append([])

    sheet.append([labels["definitions"]])
    sheet.cell(sheet.max_row, 1).font = Font(bold=True, size=13, color="174A87")
    sheet.append([labels["term"], labels["definition"]])
    _style_header(sheet[sheet.max_row])
    for item in document.get("definitions") or []:
        sheet.append([_xlsx_value(item.get("term")), _xlsx_value(item.get("definition"))])


def render_report_bundle_xlsx(document: dict[str, Any]) -> bytes:
    workbook = Workbook()
    labels = _report_labels(document)
    index_sheet = workbook.active
    index_sheet.title = labels["tocSheet"][:31]
    provider = document.get("provider") or {}
    bundle = document.get("bundle") or {}
    context = document.get("context") or {}
    selection = context.get("selection") or {}
    summary = document.get("summary") or {}
    reports = document.get("reports") or []
    project_name, scenario_name = _context_name(document)

    index_sheet.append([labels["basicReportBundle"], _xlsx_value(bundle.get("title"))])
    index_sheet.append([labels["providerStatement"], _xlsx_value(provider.get("statement"))])
    index_sheet.append([labels["project"], _xlsx_value(project_name)])
    index_sheet.append([labels["scenario"], _xlsx_value(scenario_name)])
    index_sheet.append([labels["flowRunId"], _xlsx_value(selection.get("flowRunId"))])
    index_sheet.append([labels["editingSessionId"], _xlsx_value(selection.get("editSessionId"))])
    index_sheet.append([labels["generatedAt"], _xlsx_value(context.get("generatedAt"))])
    index_sheet.append([labels["generationStatus"], _xlsx_value(summary.get("generationStatus"))])
    index_sheet.append([])
    index_sheet.append([labels["total"], labels["available"], labels["partial"], labels["noData"], labels["notApplicable"], labels["error"]])
    _style_header(index_sheet[index_sheet.max_row])
    index_sheet.append(
        [
            _xlsx_value(summary.get("totalCount")),
            _xlsx_value(summary.get("availableCount")),
            _xlsx_value(summary.get("partialCount")),
            _xlsx_value(summary.get("noDataCount")),
            _xlsx_value(summary.get("notApplicableCount")),
            _xlsx_value(summary.get("errorCount")),
        ]
    )
    index_sheet.append([])
    index_sheet.append([labels["order"], labels["code"], labels["report"], labels["group"], labels["status"], labels["dataCount"], labels["availabilityReason"], labels["sheet"]])
    _style_header(index_sheet[index_sheet.max_row])

    used_names = {index_sheet.title.casefold()}
    report_sheets: list[tuple[Any, dict[str, Any]]] = []
    for index, report_document in enumerate(reports, start=1):
        report = report_document.get("report") or {}
        availability = report_document.get("availability") or {}
        sheet_title = _bundle_sheet_title(report, used_names)
        sheet = workbook.create_sheet(sheet_title)
        report_sheets.append((sheet, report_document))
        index_sheet.append(
            [
                index,
                _xlsx_value(report.get("code")),
                _xlsx_value(report.get("title")),
                _xlsx_value(report.get("group")),
                _xlsx_value(availability.get("label") or availability.get("status")),
                _xlsx_value(availability.get("dataCount")),
                _xlsx_value(availability.get("reason")),
                sheet_title,
            ]
        )
        link_cell = index_sheet.cell(index_sheet.max_row, 8)
        link_cell.hyperlink = f"#'{sheet_title.replace(chr(39), chr(39) * 2)}'!A1"
        link_cell.style = "Hyperlink"

    for sheet, report_document in report_sheets:
        _append_bundle_report_sheet(sheet, report_document, provider)

    thin = Side(style="thin", color="D9E2EC")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
        _autosize_sheet(sheet, max_width=42)
    index_sheet.freeze_panes = "A14"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_report_pdf(html_bytes: bytes, *, batch: bool = False) -> bytes:
    if not _PDF_SCRIPT.is_file():
        raise HTTPException(status_code=503, detail="PDF renderer script is not installed.")
    node_executable = os.getenv("REPORT_PDF_NODE") or shutil.which("node")
    if not node_executable:
        raise HTTPException(status_code=503, detail="Node.js is required for PDF downloads.")
    max_html_bytes = _BATCH_PDF_MAX_HTML_BYTES if batch else _PDF_MAX_HTML_BYTES
    if len(html_bytes) > max_html_bytes:
        raise HTTPException(status_code=413, detail="The PDF report is too large to render safely.")
    creation_flags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    batch_acquired = False
    if batch:
        batch_acquired = _BATCH_PDF_SEMAPHORE.acquire(timeout=30)
        if not batch_acquired:
            raise HTTPException(status_code=503, detail="Batch PDF renderer is busy. Try again shortly.")
    try:
        acquired = _PDF_SEMAPHORE.acquire(timeout=30)
        if not acquired:
            raise HTTPException(status_code=503, detail="PDF renderer is busy. Try again shortly.")
        try:
            process_env = os.environ.copy()
            if not process_env.get("PLAYWRIGHT_BROWSERS_PATH") and _LOCAL_PLAYWRIGHT_BROWSERS.is_dir():
                process_env["PLAYWRIGHT_BROWSERS_PATH"] = "0"
            timeout_seconds = _PDF_TIMEOUT_SECONDS
            if batch:
                process_env["REPORT_PDF_MAX_INPUT_BYTES"] = str(_BATCH_PDF_MAX_HTML_BYTES)
                process_env["REPORT_PDF_WATCHDOG_MS"] = str(_BATCH_PDF_WATCHDOG_MS)
                process_env["REPORT_PDF_CONTENT_TIMEOUT_MS"] = str(_BATCH_PDF_CONTENT_TIMEOUT_MS)
                timeout_seconds = _BATCH_PDF_TIMEOUT_SECONDS
            else:
                process_env.pop("REPORT_PDF_MAX_INPUT_BYTES", None)
                process_env.pop("REPORT_PDF_WATCHDOG_MS", None)
                process_env.pop("REPORT_PDF_CONTENT_TIMEOUT_MS", None)
            try:
                result = subprocess.run(
                    [node_executable, str(_PDF_SCRIPT)],
                    input=html_bytes,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    shell=False,
                    check=False,
                    timeout=timeout_seconds,
                    creationflags=creation_flags,
                    env=process_env,
                )
            except subprocess.TimeoutExpired as error:
                raise HTTPException(status_code=504, detail="PDF rendering timed out.") from error
            if result.returncode != 0 or not result.stdout.startswith(b"%PDF-"):
                error_text = result.stderr.decode("utf-8", errors="replace")[:1000]
                logger.error("M06001 PDF rendering failed: %s", error_text or "unknown error")
                raise HTTPException(status_code=503, detail="PDF rendering failed. Check the server renderer configuration.")
            return result.stdout
        finally:
            _PDF_SEMAPHORE.release()
    finally:
        if batch_acquired:
            _BATCH_PDF_SEMAPHORE.release()


def safe_report_filename(value: Any) -> str:
    text = str(value or "report").strip()
    text = re.sub(r"[\x00-\x1f\x7f<>:\"/\\|?*]+", "_", text)
    text = re.sub(r"\s+", "_", text).strip(" ._")
    return text[:80] or "report"
