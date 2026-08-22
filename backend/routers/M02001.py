"""
@file           M02001.py
@description    File upload management API
"""

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from pydantic import BaseModel, ConfigDict
from typing import BinaryIO, Iterator, Optional
import codecs
import csv
import hashlib
import io
import json
import logging
import os
import re
import tempfile
import time
import uuid
from pathlib import Path

import oracledb
from backend.database_helper import SqlLoader, execute_query
from backend.auth_context import get_request_role_code, get_request_user_id
from backend.target_database import get_target_db_connection
from backend.paging import create_page_window, normalize_page_number, normalize_page_size

logger = logging.getLogger(__name__)
router = APIRouter()
UPLOAD_ROW_NO_COLUMN = "FILE_ROW_NO"
UPLOAD_INSERT_BATCH_SIZE = max(100, int(os.getenv("UPLOAD_INSERT_BATCH_SIZE", "2000")))
UPLOAD_INSERT_BATCH_CELL_LIMIT = max(10_000, int(os.getenv("UPLOAD_INSERT_BATCH_CELL_LIMIT", "120000")))
UPLOAD_INSERT_BATCH_BYTE_LIMIT = max(1024 * 1024, int(os.getenv("UPLOAD_INSERT_BATCH_BYTE_LIMIT", str(16 * 1024 * 1024))))
UPLOAD_COMMIT_ROW_INTERVAL = max(1000, int(os.getenv("UPLOAD_COMMIT_ROW_INTERVAL", "20000")))
UPLOAD_DIRECT_PATH_ENABLED = str(os.getenv("UPLOAD_DIRECT_PATH_ENABLED", "Y")).strip().upper() not in {"N", "NO", "FALSE", "0"}
UPLOAD_DB_CALL_TIMEOUT_MS = max(0, int(os.getenv("UPLOAD_DB_CALL_TIMEOUT_MS", "600000")))
ORACLE_COMMENT_SAFE_BYTE_LIMIT = 3_900
UPLOAD_ENCODING_SAMPLE_SIZE = 128 * 1024
UPLOAD_PREVIEW_ROW_LIMIT = 50
UPLOAD_HTTP_CHUNK_SIZE = 4 * 1024 * 1024
UPLOAD_HTTP_CHUNK_LIMIT = 8 * 1024 * 1024
UPLOAD_STAGING_MAX_AGE_SECONDS = 6 * 60 * 60
UPLOAD_STAGING_DIRECTORY = Path(tempfile.gettempdir()) / "init-data-editing-uploads"
AUTO_ENCODING_NAMES = {"", "auto", "detect", "auto-detect", "automatic"}
UPLOAD_META_READY = "READY"
UPLOAD_META_LOADING = "LOADING"


class UploadTableRequest(BaseModel):
    tableName: Optional[str] = None
    limit: Optional[int] = 100
    page: Optional[int] = 1
    model_config = ConfigDict(extra="allow")


class SqlRequest(BaseModel):
    sql: str
    limit: Optional[int] = 100
    page: Optional[int] = 1
    model_config = ConfigDict(extra="allow")


class DropTableRequest(BaseModel):
    tableName: Optional[str] = None
    model_config = ConfigDict(extra="allow")


class UploadSessionRequest(BaseModel):
    fileName: Optional[str] = None
    fileSize: int
    model_config = ConfigDict(extra="forbid")


@router.post("/upload-session")
def create_upload_session(req: UploadSessionRequest, request: Request):
    user_id = str(get_request_user_id(request))
    expected_size = max(0, int(req.fileSize or 0))
    cleanup_stale_uploads()
    UPLOAD_STAGING_DIRECTORY.mkdir(parents=True, exist_ok=True)
    upload_id = uuid.uuid4().hex
    data_path, _ = get_staging_paths(upload_id)
    data_path.touch(exist_ok=False)
    write_staged_upload_metadata(upload_id, {
        "uploadId": upload_id,
        "userId": user_id,
        "fileName": re.split(r"[\\/]", req.fileName or "uploaded-file")[-1][:255],
        "expectedSize": expected_size,
        "receivedSize": 0,
        "createdAt": time.time(),
    })
    return {
        "status": "success",
        "uploadId": upload_id,
        "chunkSize": UPLOAD_HTTP_CHUNK_SIZE,
        "receivedSize": 0,
    }


@router.post("/upload-chunk")
async def upload_file_chunk(
    request: Request,
    chunk: UploadFile = File(...),
    uploadId: str = Form(...),
    offset: int = Form(...),
):
    metadata, data_path = require_staged_upload(request, uploadId)
    expected_size = int(metadata.get("expectedSize") or 0)
    requested_offset = max(0, int(offset or 0))
    current_size = data_path.stat().st_size
    if requested_offset != current_size:
        raise HTTPException(
            status_code=409,
            detail=f"Upload offset mismatch. Expected {current_size}, received {requested_offset}.",
        )

    received = 0
    with data_path.open("r+b") as output:
        output.seek(requested_offset)
        try:
            while True:
                block = await chunk.read(256 * 1024)
                if not block:
                    break
                received += len(block)
                if received > UPLOAD_HTTP_CHUNK_LIMIT:
                    raise HTTPException(status_code=413, detail="Upload chunk is too large.")
                if requested_offset + received > expected_size:
                    raise HTTPException(status_code=400, detail="Upload exceeds the declared file size.")
                output.write(block)
        except Exception:
            output.truncate(requested_offset)
            raise

    metadata["receivedSize"] = requested_offset + received
    metadata["updatedAt"] = time.time()
    write_staged_upload_metadata(uploadId, metadata)
    return {
        "status": "success",
        "uploadId": uploadId,
        "receivedSize": metadata["receivedSize"],
        "expectedSize": expected_size,
    }


@router.post("/preview-staged")
def preview_staged_upload(
    request: Request,
    uploadId: str = Form(...),
    fileType: str = Form("csv"),
    delimiter: str = Form(","),
    fixedWidths: str = Form(""),
    hasHeader: str = Form("Y"),
    encoding: str = Form("auto"),
):
    metadata, data_path = require_completed_staged_upload(request, uploadId)
    with data_path.open("rb") as staged_file:
        staged_upload = UploadFile(file=staged_file, filename=metadata.get("fileName") or "uploaded-file")
        return preview_upload(staged_upload, fileType, delimiter, fixedWidths, hasHeader, encoding)


@router.post("/upload-staged")
def upload_staged_file_to_table(
    request: Request,
    uploadId: str = Form(...),
    fileType: str = Form("csv"),
    delimiter: str = Form(","),
    fixedWidths: str = Form(""),
    hasHeader: str = Form("Y"),
    encoding: str = Form("auto"),
    projectId: str = Form(""),
    projectCode: str = Form(""),
    tableComment: str = Form(""),
    tableNameRule: str = Form("INITUP$_{PROJECT_CODE}_FT_{TIME}"),
):
    operation_request = build_staged_operation_request(
        "UPLOAD",
        projectId,
        projectCode,
        table_name_rule=tableNameRule,
        file_type=fileType,
        delimiter=delimiter,
        fixed_widths=fixedWidths,
        has_header=hasHeader,
        encoding=encoding,
        table_comment=tableComment,
    )
    metadata, data_path, cached_result, lock_path = begin_staged_finalization(
        request,
        uploadId,
        operation_request,
    )
    if cached_result is not None:
        return cached_result
    try:
        with data_path.open("rb") as staged_file:
            staged_upload = UploadFile(file=staged_file, filename=metadata.get("fileName") or "uploaded-file")
            result = upload_file_to_table(
                request,
                staged_upload,
                fileType,
                delimiter,
                fixedWidths,
                hasHeader,
                encoding,
                projectId,
                projectCode,
                tableComment,
                tableNameRule,
                uploadId,
            )
        finalize_staged_upload(uploadId, metadata, operation_request, result)
        return result
    finally:
        release_staged_finalization_lock(lock_path)


@router.post("/reload-staged")
def reload_staged_file_into_table(
    request: Request,
    uploadId: str = Form(...),
    targetTableName: str = Form(...),
    fileType: str = Form("csv"),
    delimiter: str = Form(","),
    fixedWidths: str = Form(""),
    hasHeader: str = Form("Y"),
    encoding: str = Form("auto"),
    projectId: str = Form(""),
    projectCode: str = Form(""),
    tableComment: str = Form(""),
):
    operation_request = build_staged_operation_request(
        "RELOAD",
        projectId,
        projectCode,
        target_table_name=targetTableName,
        file_type=fileType,
        delimiter=delimiter,
        fixed_widths=fixedWidths,
        has_header=hasHeader,
        encoding=encoding,
        table_comment=tableComment,
    )
    metadata, data_path, cached_result, lock_path = begin_staged_finalization(
        request,
        uploadId,
        operation_request,
    )
    if cached_result is not None:
        return cached_result
    try:
        with data_path.open("rb") as staged_file:
            staged_upload = UploadFile(file=staged_file, filename=metadata.get("fileName") or "uploaded-file")
            result = reload_file_into_table(
                request,
                staged_upload,
                targetTableName,
                fileType,
                delimiter,
                fixedWidths,
                hasHeader,
                encoding,
                projectId,
                projectCode,
                tableComment,
                uploadId,
            )
        finalize_staged_upload(uploadId, metadata, operation_request, result)
        return result
    finally:
        release_staged_finalization_lock(lock_path)


@router.post("/preview")
def preview_upload(
    file: UploadFile = File(...),
    fileType: str = Form("csv"),
    delimiter: str = Form(","),
    fixedWidths: str = Form(""),
    hasHeader: str = Form("Y"),
    encoding: str = Form("auto")
):
    columns, rows, resolved_encoding = read_upload_preview(
        file.file,
        file.filename or "",
        fileType,
        delimiter,
        fixedWidths,
        hasHeader,
        encoding,
        UPLOAD_PREVIEW_ROW_LIMIT,
    )
    preview_columns, preview_rows = add_row_numbers_to_preview(columns, rows)
    return {
        "status": "success",
        "columns": preview_columns,
        "data": preview_rows,
        "total": len(rows),
        "detectedEncoding": resolved_encoding,
    }


@router.post("/upload")
def upload_file_to_table(
    request: Request,
    file: UploadFile = File(...),
    fileType: str = Form("csv"),
    delimiter: str = Form(","),
    fixedWidths: str = Form(""),
    hasHeader: str = Form("Y"),
    encoding: str = Form("auto"),
    projectId: str = Form(""),
    projectCode: str = Form(""),
    tableComment: str = Form(""),
    tableNameRule: str = Form("INITUP$_{PROJECT_CODE}_FT_{TIME}"),
    uploadIdempotencyKey: str = Form(""),
):
    if not str(projectId or "").strip() or not str(projectCode or "").strip():
        raise HTTPException(status_code=400, detail="Project ID and project code are required for upload.")
    user_id = get_request_user_id(request)
    require_project_access(request, projectId, projectCode)
    stream = file.file
    filename = file.filename or ""
    resolved_encoding = resolve_upload_encoding(stream, fileType, encoding)
    columns, row_width, _header_width = inspect_upload_schema(
        stream,
        filename,
        fileType,
        delimiter,
        fixedWidths,
        hasHeader,
        resolved_encoding,
    )
    if not columns:
        raise HTTPException(status_code=400, detail="No columns were detected.")

    table_name = create_upload_table_name(projectCode, tableNameRule, user_id=user_id)
    staging_table_name = create_upload_staging_table_name()
    file_size = get_upload_stream_size(stream)
    content_sha256 = compute_upload_stream_sha256(stream)
    column_specs = build_file_upload_column_specs(columns, hasHeader)
    safe_columns = [column_name for column_name, _ in column_specs]
    upload_columns = [UPLOAD_ROW_NO_COLUMN, *safe_columns]

    conn = None
    cursor = None
    published = False
    final_table_created = False
    metadata_reserved = False
    try:
        conn = get_target_db_connection(request)
        configure_upload_connection(conn)
        cursor = conn.cursor()
        ensure_upload_table_metadata(cursor)
        cleanup_stale_upload_work_tables(conn, cursor)
        for reservation_attempt in range(10):
            try:
                cursor.execute(
                    SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_RESERVE"),
                    {
                        "tableName": table_name,
                        "projectId": int(projectId),
                        "fileExtension": Path(filename).suffix.lstrip(".").upper() or None,
                        "fileName": filename or None,
                        "fileSize": file_size,
                        "uploadId": str(uploadIdempotencyKey or "").strip().lower() or None,
                        "contentSha256": content_sha256,
                    },
                )
                conn.commit()
                metadata_reserved = True
                break
            except Exception as reservation_error:
                conn.rollback()
                if "ORA-00001" not in str(reservation_error) or reservation_attempt >= 9:
                    raise
                time.sleep(0.002)
                table_name = create_upload_table_name(projectCode, tableNameRule, user_id=user_id)
        column_ddl = ", ".join([
            f'"{UPLOAD_ROW_NO_COLUMN}" NUMBER',
            *[f'"{column}" VARCHAR2(4000 BYTE)' for column in safe_columns]
        ])
        cursor.execute(f'CREATE TABLE "{staging_table_name}" ({column_ddl}) NOLOGGING')
        cursor.execute(f'COMMENT ON COLUMN "{staging_table_name}"."{UPLOAD_ROW_NO_COLUMN}" IS \'File row number\'')
        for safe_column, comment in column_specs:
            if comment:
                safe_comment = escape_and_truncate_oracle_comment(comment)
                cursor.execute(
                    f'COMMENT ON COLUMN "{staging_table_name}"."{safe_column}" IS \'{safe_comment}\''
                )
        if (tableComment or "").strip():
            safe_table_comment = escape_and_truncate_oracle_comment(tableComment.strip())
            cursor.execute(f'COMMENT ON TABLE "{staging_table_name}" IS \'{safe_table_comment}\'')

        rows = iter_upload_data_rows(
            stream,
            filename,
            fileType,
            delimiter,
            fixedWidths,
            hasHeader,
            resolved_encoding,
            row_width,
        )
        try:
            load_result = load_upload_rows(
                conn,
                cursor,
                staging_table_name,
                upload_columns,
                rows,
            )
        finally:
            close_row_iterator(rows)

        inserted_count = int(load_result["rowCount"])
        validated_count = count_upload_table_rows(cursor, staging_table_name)
        if validated_count != inserted_count:
            raise RuntimeError(
                f"Upload row validation failed. Parsed {inserted_count} row(s), but Oracle stored {validated_count} row(s)."
            )

        stats_gathered = False
        stats_message = ""
        try:
            gather_upload_table_stats(cursor, staging_table_name)
            conn.commit()
            stats_gathered = True
            stats_message = "Table statistics gathered."
        except Exception as stats_error:
            stats_message = f"Table uploaded, but statistics gather failed: {stats_error}"
            logger.warning("M02001 statistics gather failed for %s: %s", staging_table_name, stats_error)

        cursor.execute(
            SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_MERGE"),
            {
                "tableName": table_name,
                "projectId": int(projectId),
                "fileExtension": Path(filename).suffix.lstrip(".").upper() or None,
                "fileName": filename or None,
                "fileSize": file_size,
                "uploadId": str(uploadIdempotencyKey or "").strip().lower() or None,
                "rowCount": inserted_count,
                "contentSha256": content_sha256,
                "loadStatus": UPLOAD_META_LOADING,
            },
        )
        conn.commit()
        cursor.execute(f'ALTER TABLE "{staging_table_name}" LOGGING')
        cursor.execute(f'ALTER TABLE "{staging_table_name}" RENAME TO "{table_name}"')
        final_table_created = True
        staging_table_name = ""
        cursor.execute(
            SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_MERGE"),
            {
                "tableName": table_name,
                "projectId": int(projectId),
                "fileExtension": Path(filename).suffix.lstrip(".").upper() or None,
                "fileName": filename or None,
                "fileSize": file_size,
                "uploadId": str(uploadIdempotencyKey or "").strip().lower() or None,
                "rowCount": inserted_count,
                "contentSha256": content_sha256,
                "loadStatus": UPLOAD_META_READY,
            },
        )
        conn.commit()
        published = True
        return {
            "status": "success",
            "message": "File uploaded.",
            "tableName": table_name,
            "columns": upload_columns,
            "rowCount": inserted_count,
            "detectedEncoding": resolved_encoding,
            "contentSha256": content_sha256,
            "loadMode": load_result["loadMode"],
            "insertBatchCount": load_result["batchCount"],
            "commitCount": load_result["commitCount"],
            "statsGathered": stats_gathered,
            "statsMessage": stats_message
        }
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"M02001 upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if cursor and not published:
            cleanup_upload_work_table(
                conn,
                cursor,
                staging_table_name or (table_name if final_table_created else ""),
            )
            if metadata_reserved:
                cleanup_upload_table_metadata(conn, cursor, table_name)
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/reload")
def reload_file_into_table(
    request: Request,
    file: UploadFile = File(...),
    targetTableName: str = Form(...),
    fileType: str = Form("csv"),
    delimiter: str = Form(","),
    fixedWidths: str = Form(""),
    hasHeader: str = Form("Y"),
    encoding: str = Form("auto"),
    projectId: str = Form(""),
    projectCode: str = Form(""),
    tableComment: str = Form(""),
    uploadIdempotencyKey: str = Form(""),
):
    if not str(projectId or "").strip() or not str(projectCode or "").strip():
        raise HTTPException(status_code=400, detail="Project ID and project code are required for reload.")
    require_project_access(request, projectId, projectCode)
    table_name = require_upload_table(targetTableName)
    stream = file.file
    filename = file.filename or ""
    resolved_encoding = resolve_upload_encoding(stream, fileType, encoding)
    columns, row_width, header_width = inspect_upload_schema(
        stream,
        filename,
        fileType,
        delimiter,
        fixedWidths,
        hasHeader,
        resolved_encoding,
    )
    if not columns:
        raise HTTPException(status_code=400, detail="No columns were detected.")

    file_size = get_upload_stream_size(stream)
    content_sha256 = compute_upload_stream_sha256(stream)
    column_specs = build_file_upload_column_specs(columns, hasHeader)
    staging_table_name = create_upload_staging_table_name()
    conn = None
    cursor = None
    try:
        conn = get_target_db_connection(request)
        configure_upload_connection(conn)
        cursor = conn.cursor()
        ensure_upload_table_metadata(cursor)
        cleanup_stale_upload_work_tables(conn, cursor)

        base_prefix = create_upload_table_prefix(projectCode)
        table_result = execute_query(
            conn,
            "M02001_UPLOAD_TABLE_TREE",
            {"tablePrefix": base_prefix, "projectId": int(projectId)},
        )
        allowed_tables = {
            str(row.get("TABLE_NAME") or "").upper()
            for row in table_result.get("data", [])
        } if table_result.get("status") == "success" else set()
        if table_name not in allowed_tables:
            raise HTTPException(
                status_code=404,
                detail="The selected reload table does not belong to the current project or no longer exists.",
            )

        column_result = execute_query(conn, "M02001_UPLOAD_TABLE_COLUMNS", {"tableName": table_name})
        target_columns = column_result.get("data", []) if column_result.get("status") == "success" else []
        if not target_columns:
            raise HTTPException(status_code=404, detail="The selected reload table has no readable columns.")

        source_width = header_width if str(hasHeader or "Y").upper() == "Y" else row_width
        target_column_names = validate_reload_column_layout(target_columns, source_width)
        data_column_names = target_column_names[1:]
        quoted_table = f'"{table_name}"'
        quoted_columns = ", ".join(f'"{column_name}"' for column_name in target_column_names)
        cursor.execute(
            f'CREATE TABLE "{staging_table_name}" NOLOGGING AS '
            f'SELECT /*+ NO_PARALLEL */ {quoted_columns} FROM {quoted_table} WHERE 1=0'
        )

        rows = iter_upload_data_rows(
            stream,
            filename,
            fileType,
            delimiter,
            fixedWidths,
            hasHeader,
            resolved_encoding,
            row_width,
        )
        try:
            load_result = load_upload_rows(
                conn,
                cursor,
                staging_table_name,
                target_column_names,
                rows,
            )
        finally:
            close_row_iterator(rows)
        inserted_count = int(load_result["rowCount"])
        validated_count = count_upload_table_rows(cursor, staging_table_name)
        if validated_count != inserted_count:
            raise RuntimeError(
                f"Reload row validation failed. Parsed {inserted_count} row(s), "
                f"but Oracle stored {validated_count} row(s)."
            )

        cursor.execute(f"DELETE FROM {quoted_table}")
        cursor.execute(
            f'INSERT /*+ NO_PARALLEL */ INTO {quoted_table} ({quoted_columns}) '
            f'SELECT /*+ NO_PARALLEL */ {quoted_columns} FROM "{staging_table_name}"'
        )
        copied_count = int(cursor.rowcount or 0)
        if copied_count >= 0 and copied_count != inserted_count:
            raise RuntimeError(
                f"Reload copy validation failed. Expected {inserted_count} row(s), copied {copied_count} row(s)."
            )

        cursor.execute(
            SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_MERGE"),
            {
                "tableName": table_name,
                "projectId": int(projectId),
                "fileExtension": Path(filename).suffix.lstrip(".").upper() or None,
                "fileName": filename or None,
                "fileSize": file_size,
                "uploadId": str(uploadIdempotencyKey or "").strip().lower() or None,
                "rowCount": inserted_count,
                "contentSha256": content_sha256,
                "loadStatus": UPLOAD_META_READY,
            },
        )
        conn.commit()

        metadata_messages = []
        try:
            for index, (unused_column_name, comment) in enumerate(column_specs):
                if not comment:
                    continue
                target_column = data_column_names[index]
                safe_comment = escape_and_truncate_oracle_comment(comment)
                cursor.execute(
                    f'COMMENT ON COLUMN "{table_name}"."{target_column}" IS \'{safe_comment}\''
                )
            if (tableComment or "").strip():
                safe_table_comment = escape_and_truncate_oracle_comment(tableComment.strip())
                cursor.execute(f'COMMENT ON TABLE "{table_name}" IS \'{safe_table_comment}\'')
        except Exception as metadata_error:
            metadata_messages.append(f"Comment update failed: {metadata_error}")
            logger.warning("M02001 reload comment update failed for %s: %s", table_name, metadata_error)

        stats_gathered = False
        try:
            gather_upload_table_stats(cursor, table_name)
            stats_gathered = True
        except Exception as stats_error:
            metadata_messages.append(f"Statistics gather failed: {stats_error}")
            logger.warning("M02001 reload statistics gather failed for %s: %s", table_name, stats_error)
        conn.commit()
        return {
            "status": "success",
            "message": "File reloaded.",
            "tableName": table_name,
            "columns": target_column_names,
            "rowCount": inserted_count,
            "detectedEncoding": resolved_encoding,
            "contentSha256": content_sha256,
            "loadMode": load_result["loadMode"],
            "insertBatchCount": load_result["batchCount"],
            "commitCount": load_result["commitCount"],
            "statsGathered": stats_gathered,
            "statsMessage": " ".join(metadata_messages),
        }
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as error:
        if conn:
            conn.rollback()
        logger.error("M02001 reload failed for %s: %s", table_name, error)
        raise HTTPException(status_code=500, detail=str(error)) from error
    finally:
        if cursor:
            cleanup_upload_work_table(conn, cursor, staging_table_name)
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/drop-table")
def drop_upload_table(req: DropTableRequest, request: Request):
    table_name = require_upload_table(req.tableName)
    conn = None
    cursor = None
    try:
        conn = get_target_db_connection(request)
        cursor = conn.cursor()
        ensure_upload_table_metadata(cursor)
        cursor.execute(f'DROP TABLE "{table_name}" PURGE')
        cursor.execute(SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_DELETE"), {"tableName": table_name})
        conn.commit()
        return {
            "status": "success",
            "message": "Upload table dropped.",
            "tableName": table_name
        }
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"M02001 drop table failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/upload-table-tree")
def get_upload_table_tree(
    request: Request,
    projectId: str = "",
    projectCode: str = "",
    tablePrefix: str = "",
):
    if not str(projectId or "").strip() or not str(projectCode or "").strip():
        raise HTTPException(status_code=400, detail="Project ID and project code are required.")
    require_project_access(request, projectId, projectCode)
    base_prefix = create_upload_table_prefix(projectCode)
    table_prefix = normalize_upload_table_search_prefix(tablePrefix, base_prefix)
    conn = None
    cursor = None
    try:
        conn = get_target_db_connection(request)
        cursor = conn.cursor()
        ensure_upload_table_metadata(cursor)
        conn.commit()
        result = execute_query(
            conn,
            "M02001_UPLOAD_TABLE_TREE",
            {"tablePrefix": table_prefix, "projectId": int(projectId)},
        )
        if result.get("status") != "success":
            raise HTTPException(status_code=500, detail=result.get("detail") or result.get("message") or "Upload table tree query failed.")
        return {
            "status": "success",
            "data": result.get("data", []),
            "columns": result.get("columns", []),
            "total": result.get("total", 0),
            "tablePrefix": table_prefix
        }
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/columns")
def get_upload_columns(req: UploadTableRequest, request: Request):
    table_name = require_read_table(req.tableName)
    conn = None
    try:
        conn = get_target_db_connection(request)
        result = execute_query(conn, "M02001_UPLOAD_TABLE_COLUMNS", {"tableName": table_name})
        if result.get("status") != "success":
            raise HTTPException(status_code=500, detail=result.get("detail") or result.get("message") or "Column query failed.")
        return {
            "status": "success",
            "data": result.get("data", []),
            "columns": result.get("columns", []),
            "total": result.get("total", 0)
        }
    finally:
        if conn:
            conn.close()


@router.post("/data")
def get_upload_data(req: UploadTableRequest, request: Request):
    table_name = require_read_table(req.tableName)
    limit = normalize_page_size(req.limit)
    page = normalize_page_number(req.page)
    conn = None
    try:
        conn = get_target_db_connection(request)
        table_object = f'"{table_name}"'
        count_result = execute_query(conn, "M02001_UPLOAD_TABLE_DATA_COUNT", {"dynamicTable": table_object})
        total_rows = count_result.get("data", [])
        total = int(total_rows[0].get("TOTAL_COUNT") or 0) if total_rows else 0
        page_window = create_page_window(page, limit, total)
        result = execute_query(conn, "M02001_UPLOAD_TABLE_DATA_PAGE", {"dynamicTable": table_object, "offset": page_window.offset, "limit": page_window.page_size})
        if result.get("status") != "success":
            raise HTTPException(status_code=500, detail=result.get("detail") or result.get("message") or "Data query failed.")
        response = {
            "status": "success",
            "data": result.get("data", []),
            "columns": result.get("columns", []),
            "total": total
        }
        response.update(page_window.response_metadata())
        return response
    finally:
        if conn:
            conn.close()


@router.post("/sql")
def execute_sql(req: SqlRequest, request: Request):
    sql = normalize_select_sql(req.sql)
    limit = normalize_page_size(req.limit)
    page = normalize_page_number(req.page)
    conn = None
    try:
        conn = get_target_db_connection(request)
        count_result = execute_query(conn, "M02001_SQL_WORKSHEET_COUNT", {"dynamicSql": sql})
        total_rows = count_result.get("data", [])
        total = int(total_rows[0].get("TOTAL_COUNT") or 0) if total_rows else 0
        page_window = create_page_window(page, limit, total)
        result = execute_query(conn, "M02001_SQL_WORKSHEET_PAGE", {"dynamicSql": sql, "offset": page_window.offset, "limit": page_window.page_size})
        if result.get("status") != "success":
            raise HTTPException(status_code=500, detail=result.get("detail") or result.get("message") or "SQL execution failed.")
        response = {
            "status": "success",
            "data": result.get("data", []),
            "columns": result.get("columns", []),
            "total": total
        }
        response.update(page_window.response_metadata())
        return response
    finally:
        if conn:
            conn.close()


def get_staging_paths(upload_id: str):
    normalized_id = str(upload_id or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{32}", normalized_id):
        raise HTTPException(status_code=400, detail="Invalid upload session ID.")
    return (
        UPLOAD_STAGING_DIRECTORY / f"{normalized_id}.upload",
        UPLOAD_STAGING_DIRECTORY / f"{normalized_id}.json",
    )


def write_staged_upload_metadata(upload_id: str, metadata: dict):
    UPLOAD_STAGING_DIRECTORY.mkdir(parents=True, exist_ok=True)
    _, metadata_path = get_staging_paths(upload_id)
    temporary_path = metadata_path.with_suffix(".json.tmp")
    with temporary_path.open("w", encoding="utf-8") as metadata_file:
        json.dump(metadata, metadata_file, ensure_ascii=False)
    os.replace(temporary_path, metadata_path)


def read_staged_upload_metadata(request: Request, upload_id: str):
    _data_path, metadata_path = get_staging_paths(upload_id)
    try:
        with metadata_path.open("r", encoding="utf-8") as metadata_file:
            metadata = json.load(metadata_file)
    except (FileNotFoundError, json.JSONDecodeError) as error:
        raise HTTPException(status_code=404, detail="Upload session was not found or has expired.") from error
    if float(metadata.get("createdAt") or 0) < time.time() - UPLOAD_STAGING_MAX_AGE_SECONDS:
        discard_staged_upload(upload_id)
        raise HTTPException(status_code=404, detail="Upload session was not found or has expired.")
    if str(metadata.get("userId") or "") != str(get_request_user_id(request)):
        raise HTTPException(status_code=404, detail="Upload session was not found or has expired.")
    return metadata


def require_staged_upload(request: Request, upload_id: str):
    data_path, metadata_path = get_staging_paths(upload_id)
    metadata = read_staged_upload_metadata(request, upload_id)
    if not data_path.is_file():
        raise HTTPException(status_code=404, detail="Upload session data was not found.")
    return metadata, data_path


def require_completed_staged_upload(request: Request, upload_id: str):
    metadata, data_path = require_staged_upload(request, upload_id)
    expected_size = int(metadata.get("expectedSize") or 0)
    received_size = data_path.stat().st_size
    if received_size != expected_size:
        raise HTTPException(
            status_code=409,
            detail=f"Upload is incomplete. Expected {expected_size} bytes, received {received_size} bytes.",
        )
    return metadata, data_path


def discard_staged_upload(upload_id: str):
    data_path, metadata_path = get_staging_paths(upload_id)
    for path in (
        data_path,
        metadata_path,
        metadata_path.with_suffix(".json.tmp"),
        metadata_path.with_suffix(".lock"),
    ):
        try:
            path.unlink(missing_ok=True)
        except OSError:
            logger.warning("M02001 staged upload cleanup failed for %s", path.name)


def build_staged_operation_request(
    operation: str,
    project_id: str,
    project_code: str,
    table_name_rule: str = "",
    target_table_name: str = "",
    file_type: str = "",
    delimiter: str = "",
    fixed_widths: str = "",
    has_header: str = "",
    encoding: str = "",
    table_comment: str = "",
):
    return {
        "operation": str(operation or "").strip().upper(),
        "projectId": str(project_id or "").strip(),
        "projectCode": str(project_code or "").strip().upper(),
        "tableNameRule": str(table_name_rule or "").strip(),
        "targetTableName": str(target_table_name or "").strip().upper(),
        "fileType": str(file_type or "").strip().lower(),
        "delimiter": str(delimiter or ""),
        "fixedWidths": str(fixed_widths or "").strip(),
        "hasHeader": str(has_header or "").strip().upper(),
        "encoding": str(encoding or "").strip().lower(),
        "tableComment": str(table_comment or ""),
    }


def begin_staged_finalization(request: Request, upload_id: str, operation_request: dict):
    metadata = read_staged_upload_metadata(request, upload_id)
    cached_result = metadata.get("finalResult")
    if cached_result is not None:
        if metadata.get("finalRequest") != operation_request:
            raise HTTPException(
                status_code=409,
                detail="This upload session was already finalized for a different operation.",
            )
        return metadata, None, cached_result, None

    _data_path, metadata_path = get_staging_paths(upload_id)
    lock_path = metadata_path.with_suffix(".lock")
    try:
        descriptor = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.close(descriptor)
    except FileExistsError as error:
        raise HTTPException(status_code=409, detail="This upload session is already being finalized.") from error

    try:
        metadata = read_staged_upload_metadata(request, upload_id)
        cached_result = metadata.get("finalResult")
        if cached_result is not None:
            if metadata.get("finalRequest") != operation_request:
                raise HTTPException(
                    status_code=409,
                    detail="This upload session was already finalized for a different operation.",
                )
            release_staged_finalization_lock(lock_path)
            return metadata, None, cached_result, None
        metadata, data_path = require_completed_staged_upload(request, upload_id)
        return metadata, data_path, None, lock_path
    except Exception:
        release_staged_finalization_lock(lock_path)
        raise


def finalize_staged_upload(upload_id: str, metadata: dict, operation_request: dict, result: dict):
    metadata["finalRequest"] = operation_request
    metadata["finalResult"] = result
    metadata["finalizedAt"] = time.time()
    metadata["updatedAt"] = metadata["finalizedAt"]
    write_staged_upload_metadata(upload_id, metadata)
    data_path, _metadata_path = get_staging_paths(upload_id)
    try:
        data_path.unlink(missing_ok=True)
    except OSError:
        logger.warning("M02001 finalized upload data cleanup failed for %s", data_path.name)


def release_staged_finalization_lock(lock_path) -> None:
    if not lock_path:
        return
    try:
        lock_path.unlink(missing_ok=True)
    except OSError:
        logger.warning("M02001 staged upload lock cleanup failed for %s", lock_path.name)


def ensure_upload_table_metadata(cursor) -> None:
    cursor.execute(SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_EXISTS"))
    if int((cursor.fetchone() or [0])[0] or 0) <= 0:
        try:
            cursor.execute(SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_CREATE"))
        except Exception as error:
            if "ORA-00955" not in str(error):
                raise
    for column_key in (
        "PROJECT",
        "UPLOAD_ID",
        "ROW_COUNT",
        "CONTENT_SHA256",
        "LOAD_STATUS",
    ):
        cursor.execute(SqlLoader.get_sql(f"M02001_UPLOAD_TABLE_META_{column_key}_EXISTS"))
        if int((cursor.fetchone() or [0])[0] or 0) > 0:
            continue
        try:
            cursor.execute(SqlLoader.get_sql(f"M02001_UPLOAD_TABLE_META_{column_key}_ADD"))
        except Exception as error:
            if "ORA-01430" not in str(error):
                raise


def get_upload_stream_size(stream: BinaryIO) -> int:
    try:
        position = stream.tell()
        stream.seek(0, os.SEEK_END)
        size = stream.tell()
        stream.seek(position)
        return max(0, int(size))
    except Exception:
        return 0


def configure_upload_connection(conn) -> None:
    try:
        conn.call_timeout = UPLOAD_DB_CALL_TIMEOUT_MS
    except Exception:
        logger.debug("M02001 upload call timeout could not be configured.", exc_info=True)


def compute_upload_stream_sha256(stream: BinaryIO) -> str:
    digest = hashlib.sha256()
    stream.seek(0)
    while True:
        block = stream.read(1024 * 1024)
        if not block:
            break
        digest.update(block)
    stream.seek(0)
    return digest.hexdigest()


def create_upload_staging_table_name() -> str:
    return f"INITSTG$_{uuid.uuid4().hex.upper()}"


def get_upload_insert_batch_size(column_count: int) -> int:
    return max(
        1,
        min(
            UPLOAD_INSERT_BATCH_SIZE,
            UPLOAD_INSERT_BATCH_CELL_LIMIT // max(int(column_count or 0), 1),
        ),
    )


def prepare_upload_bind_row(row_number: int, row: list[str]):
    estimated_bytes = 16
    normalized = []
    for column_index, value in enumerate(row, start=1):
        text_value = stringify_cell(value)
        byte_length = len(text_value.encode("utf-8"))
        if byte_length > 4000:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"File data row {row_number}, column {column_index} exceeds "
                    f"the Oracle VARCHAR2(4000 BYTE) limit ({byte_length} bytes)."
                ),
            )
        estimated_bytes += byte_length + 8
        normalized.append(text_value)
    return (row_number, *normalized), estimated_bytes


def iter_upload_insert_batches(rows: Iterator[list[str]], column_count: int):
    max_rows = get_upload_insert_batch_size(column_count)
    batch_rows = []
    batch_bytes = 0
    for row_number, row in enumerate(rows, start=1):
        bind_row, row_bytes = prepare_upload_bind_row(row_number, row)
        if batch_rows and (
            len(batch_rows) >= max_rows
            or batch_bytes + row_bytes > UPLOAD_INSERT_BATCH_BYTE_LIMIT
        ):
            yield batch_rows
            batch_rows = []
            batch_bytes = 0
        batch_rows.append(bind_row)
        batch_bytes += row_bytes
    if batch_rows:
        yield batch_rows


def load_upload_rows(conn, cursor, table_name: str, upload_columns: list[str], rows: Iterator[list[str]]):
    table_name = require_generated_upload_work_table(table_name)
    column_sql = ", ".join(f'"{column}"' for column in upload_columns)
    bind_sql = ", ".join(f":{index + 1}" for index in range(len(upload_columns)))
    insert_sql = f'INSERT /*+ NO_PARALLEL */ INTO "{table_name}" ({column_sql}) VALUES ({bind_sql})'
    use_direct_path = can_use_upload_direct_path(conn)
    schema_name = get_current_schema_name(cursor) if use_direct_path else ""
    inserted_count = 0
    pending_commit_rows = 0
    batch_count = 0
    commit_count = 0
    for batch_rows in iter_upload_insert_batches(rows, len(upload_columns)):
        if use_direct_path:
            conn.direct_path_load(
                schema_name,
                table_name,
                upload_columns,
                batch_rows,
                batch_size=len(batch_rows),
            )
        else:
            cursor.executemany(insert_sql, batch_rows)
        batch_size = len(batch_rows)
        inserted_count += batch_size
        pending_commit_rows += batch_size
        batch_count += 1
        if pending_commit_rows >= UPLOAD_COMMIT_ROW_INTERVAL:
            conn.commit()
            commit_count += 1
            pending_commit_rows = 0
    if pending_commit_rows or not batch_count:
        conn.commit()
        commit_count += 1
    return {
        "rowCount": inserted_count,
        "batchCount": batch_count,
        "commitCount": commit_count,
        "loadMode": "DIRECT_PATH" if use_direct_path else "ARRAY_DML",
    }


def can_use_upload_direct_path(conn) -> bool:
    return bool(
        UPLOAD_DIRECT_PATH_ENABLED
        and oracledb.is_thin_mode()
        and callable(getattr(conn, "direct_path_load", None))
    )


def get_current_schema_name(cursor) -> str:
    cursor.execute(SqlLoader.get_sql("M02001_CURRENT_SCHEMA"))
    schema_name = str((cursor.fetchone() or [""])[0] or "").strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z0-9_$#]{0,127}", schema_name):
        raise RuntimeError("Oracle current schema could not be resolved for direct path load.")
    return schema_name


def count_upload_table_rows(cursor, table_name: str) -> int:
    table_name = require_generated_upload_work_table(table_name)
    cursor.execute(f'SELECT /*+ NO_PARALLEL */ COUNT(*) FROM "{table_name}"')
    return int((cursor.fetchone() or [0])[0] or 0)


def require_generated_upload_work_table(table_name: str) -> str:
    name = str(table_name or "").strip().upper()
    if not re.fullmatch(r"(?:INITSTG\$_[A-F0-9]{32}|INITUP\$_[A-Z0-9_$#_]*[0-9]{13})", name):
        raise ValueError("Invalid generated upload work table name.")
    return name


def cleanup_upload_work_table(conn, cursor, table_name: str) -> None:
    if not conn or not cursor or not table_name:
        return
    try:
        safe_name = require_generated_upload_work_table(table_name)
        cursor.execute(f'DROP TABLE "{safe_name}" PURGE')
        conn.commit()
    except Exception as cleanup_error:
        try:
            conn.rollback()
        except Exception:
            pass
        if "ORA-00942" not in str(cleanup_error):
            logger.warning("M02001 upload work table cleanup failed for %s: %s", table_name, cleanup_error)


def cleanup_stale_upload_work_tables(conn, cursor) -> None:
    cursor.execute(
        SqlLoader.get_sql("M02001_STALE_UPLOAD_STAGE_TABLES"),
        {"maxAgeSeconds": UPLOAD_STAGING_MAX_AGE_SECONDS},
    )
    stale_stage_tables = [str(row[0] or "").strip().upper() for row in cursor.fetchall()]
    cursor.execute(
        SqlLoader.get_sql("M02001_STALE_LOADING_UPLOAD_TABLES"),
        {"maxAgeSeconds": UPLOAD_STAGING_MAX_AGE_SECONDS},
    )
    stale_loading_tables = [str(row[0] or "").strip().upper() for row in cursor.fetchall()]
    for table_name in [*stale_stage_tables, *stale_loading_tables]:
        try:
            safe_name = require_generated_upload_work_table(table_name)
            cursor.execute(f'DROP TABLE "{safe_name}" PURGE')
        except Exception as cleanup_error:
            if "ORA-00942" not in str(cleanup_error):
                logger.warning("M02001 stale Oracle upload stage cleanup failed for %s: %s", table_name, cleanup_error)
    for table_name in stale_loading_tables:
        cursor.execute(SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_DELETE"), {"tableName": table_name})
    if stale_stage_tables or stale_loading_tables:
        conn.commit()


def cleanup_upload_table_metadata(conn, cursor, table_name: str) -> None:
    if not conn or not cursor or not table_name:
        return
    try:
        safe_name = require_upload_table(table_name)
        cursor.execute(SqlLoader.get_sql("M02001_UPLOAD_TABLE_META_DELETE"), {"tableName": safe_name})
        conn.commit()
    except Exception as cleanup_error:
        try:
            conn.rollback()
        except Exception:
            pass
        logger.warning("M02001 upload metadata cleanup failed for %s: %s", table_name, cleanup_error)


def cleanup_stale_uploads():
    if not UPLOAD_STAGING_DIRECTORY.is_dir():
        return
    expiration_time = time.time() - UPLOAD_STAGING_MAX_AGE_SECONDS
    for path in UPLOAD_STAGING_DIRECTORY.iterdir():
        try:
            if path.is_file() and path.stat().st_mtime < expiration_time:
                path.unlink(missing_ok=True)
        except OSError:
            logger.warning("M02001 stale upload cleanup failed for %s", path.name)


def resolve_upload_encoding(stream: BinaryIO, file_type: str, requested_encoding: str) -> Optional[str]:
    if (file_type or "").strip().lower() == "excel":
        return None

    requested = (requested_encoding or "").strip()
    if requested.lower() not in AUTO_ENCODING_NAMES:
        try:
            return codecs.lookup(requested).name
        except LookupError as error:
            raise HTTPException(status_code=400, detail=f"Unsupported encoding: {requested}") from error

    stream.seek(0)
    sample = stream.read(UPLOAD_ENCODING_SAMPLE_SIZE)
    stream.seek(0)
    if not sample:
        return "utf-8-sig"

    bom_encodings = (
        (codecs.BOM_UTF32_LE, "utf-32"),
        (codecs.BOM_UTF32_BE, "utf-32"),
        (codecs.BOM_UTF8, "utf-8-sig"),
        (codecs.BOM_UTF16_LE, "utf-16"),
        (codecs.BOM_UTF16_BE, "utf-16"),
    )
    for bom, encoding_name in bom_encodings:
        if sample.startswith(bom):
            return encoding_name

    utf16_encoding = detect_utf16_without_bom(sample)
    if utf16_encoding:
        return utf16_encoding

    for encoding_name in ("utf-8", "cp949", "shift_jis", "big5", "windows-1252"):
        if can_decode_sample(sample, encoding_name):
            return encoding_name
    return "latin-1"


def detect_utf16_without_bom(sample: bytes) -> Optional[str]:
    if len(sample) < 4:
        return None
    even_bytes = sample[0::2]
    odd_bytes = sample[1::2]
    even_null_ratio = even_bytes.count(0) / max(len(even_bytes), 1)
    odd_null_ratio = odd_bytes.count(0) / max(len(odd_bytes), 1)
    if odd_null_ratio >= 0.3 and even_null_ratio <= 0.05:
        return "utf-16-le"
    if even_null_ratio >= 0.3 and odd_null_ratio <= 0.05:
        return "utf-16-be"
    return None


def can_decode_sample(sample: bytes, encoding: str) -> bool:
    try:
        decoder = codecs.getincrementaldecoder(encoding)(errors="strict")
        decoder.decode(sample, final=False)
        return True
    except UnicodeDecodeError:
        return False


def read_upload_preview(
    stream: BinaryIO,
    filename: str,
    file_type: str,
    delimiter: str,
    fixed_widths: str,
    has_header: str,
    encoding: str,
    preview_limit: int,
):
    resolved_encoding = resolve_upload_encoding(stream, file_type, encoding)
    use_header = str(has_header or "Y").upper() == "Y"
    columns = []
    rows = []
    width = 0
    raw_rows = iter_upload_raw_rows(
        stream,
        filename,
        file_type,
        delimiter,
        fixed_widths,
        resolved_encoding,
    )
    try:
        for raw_row in raw_rows:
            if not is_non_empty_row(raw_row):
                continue
            width = max(width, len(raw_row))
            if use_header and not columns:
                columns = build_header_columns(raw_row)
                continue
            rows.append(raw_row)
            if len(rows) >= preview_limit:
                break
    finally:
        close_row_iterator(raw_rows)

    if not columns and not rows:
        return [], [], resolved_encoding
    if not use_header:
        columns = build_default_columns(width)
    width = max(width, len(columns))
    columns = extend_columns(columns, width)
    normalized_rows = [normalize_upload_row(row, width) for row in rows]
    return columns, normalized_rows, resolved_encoding


def inspect_upload_schema(
    stream: BinaryIO,
    filename: str,
    file_type: str,
    delimiter: str,
    fixed_widths: str,
    has_header: str,
    resolved_encoding: Optional[str],
):
    """Read only the first non-empty record and validate later records while loading."""
    use_header = str(has_header or "Y").upper() == "Y"
    raw_rows = iter_upload_raw_rows(
        stream,
        filename,
        file_type,
        delimiter,
        fixed_widths,
        resolved_encoding,
    )
    try:
        first_row = next((row for row in raw_rows if is_non_empty_row(row)), None)
    finally:
        close_row_iterator(raw_rows)

    if first_row is None:
        return [], 0, 0
    width = len(first_row)
    columns = build_header_columns(first_row) if use_header else build_default_columns(width)
    return columns, width, width if use_header else 0


def inspect_upload_stream(
    stream: BinaryIO,
    filename: str,
    file_type: str,
    delimiter: str,
    fixed_widths: str,
    has_header: str,
    resolved_encoding: Optional[str],
):
    use_header = str(has_header or "Y").upper() == "Y"
    columns = []
    width = 0
    header_width = 0
    has_data = False
    raw_rows = iter_upload_raw_rows(
        stream,
        filename,
        file_type,
        delimiter,
        fixed_widths,
        resolved_encoding,
    )
    try:
        for raw_row in raw_rows:
            if not is_non_empty_row(raw_row):
                continue
            width = max(width, len(raw_row))
            if use_header and not columns:
                columns = build_header_columns(raw_row)
                header_width = len(raw_row)
                continue
            has_data = True
    finally:
        close_row_iterator(raw_rows)

    if not columns and not has_data:
        return [], 0, 0
    if not use_header:
        columns = build_default_columns(width)
    width = max(width, len(columns))
    return extend_columns(columns, width), width, header_width


def iter_upload_data_rows(
    stream: BinaryIO,
    filename: str,
    file_type: str,
    delimiter: str,
    fixed_widths: str,
    has_header: str,
    resolved_encoding: Optional[str],
    width: int,
) -> Iterator[list[str]]:
    use_header = str(has_header or "Y").upper() == "Y"
    header_skipped = False
    source_row_number = 0
    raw_rows = iter_upload_raw_rows(
        stream,
        filename,
        file_type,
        delimiter,
        fixed_widths,
        resolved_encoding,
    )
    try:
        for raw_row in raw_rows:
            source_row_number += 1
            if not is_non_empty_row(raw_row):
                continue
            if use_header and not header_skipped:
                header_skipped = True
                continue
            if len(raw_row) > width:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"File row {source_row_number} has {len(raw_row)} column(s), "
                        f"which exceeds the expected {width} column(s)."
                    ),
                )
            yield normalize_upload_row(raw_row, width)
    finally:
        close_row_iterator(raw_rows)


def iter_upload_raw_rows(
    stream: BinaryIO,
    filename: str,
    file_type: str,
    delimiter: str,
    fixed_widths: str,
    resolved_encoding: Optional[str],
) -> Iterator[list]:
    normalized_type = (file_type or "csv").strip().lower()
    if normalized_type == "excel":
        yield from iter_excel_rows(stream, filename)
        return
    if normalized_type == "fixed":
        yield from iter_fixed_rows(stream, fixed_widths, resolved_encoding or "utf-8-sig")
        return
    actual_delimiter = "\t" if normalized_type == "tsv" else (delimiter or ",")
    yield from iter_delimited_rows(stream, actual_delimiter, resolved_encoding or "utf-8-sig")


def iter_delimited_rows(stream: BinaryIO, delimiter: str, encoding: str) -> Iterator[list[str]]:
    stream.seek(0)
    text_stream = io.TextIOWrapper(stream, encoding=encoding, errors="replace", newline="")
    try:
        if len(delimiter) == 1:
            yield from csv.reader(text_stream, delimiter=delimiter)
            return
        for line in text_stream:
            yield line.rstrip("\r\n").split(delimiter)
    finally:
        detach_text_stream(text_stream)


def iter_fixed_rows(stream: BinaryIO, fixed_widths: str, encoding: str) -> Iterator[list[str]]:
    try:
        widths = [int(value.strip()) for value in (fixed_widths or "").split(",") if value.strip()]
    except ValueError as error:
        raise HTTPException(status_code=400, detail="Fixed widths must be comma-separated integers.") from error
    if not widths or any(width <= 0 for width in widths):
        raise HTTPException(status_code=400, detail="Fixed widths are required and must be positive integers.")

    stream.seek(0)
    text_stream = io.TextIOWrapper(stream, encoding=encoding, errors="replace", newline="")
    try:
        for line in text_stream:
            line = line.rstrip("\r\n")
            start = 0
            row = []
            for width in widths:
                row.append(line[start:start + width].strip())
                start += width
            yield row
    finally:
        detach_text_stream(text_stream)


def iter_excel_rows(stream: BinaryIO, filename: str) -> Iterator[list[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as error:
        raise HTTPException(status_code=500, detail="Excel upload requires openpyxl.") from error

    stream.seek(0)
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Excel file could not be read: {filename or 'uploaded file'}") from error
    try:
        for row in workbook.active.iter_rows(values_only=True):
            yield [stringify_cell(cell) for cell in row]
    finally:
        workbook.close()
        stream.seek(0)


def detach_text_stream(text_stream: io.TextIOWrapper):
    try:
        text_stream.detach()
    except (ValueError, OSError):
        pass


def close_row_iterator(rows):
    close = getattr(rows, "close", None)
    if callable(close):
        close()


def is_non_empty_row(row) -> bool:
    return any(stringify_cell(cell).strip() for cell in row)


def build_header_columns(row):
    return [stringify_cell(cell).strip() or f"COL{index + 1:03d}" for index, cell in enumerate(row)]


def build_default_columns(width):
    return [f"COL{index + 1:03d}" for index in range(width)]


def extend_columns(columns, width):
    return [*columns, *build_default_columns(width)[len(columns):width]]


def normalize_upload_row(row, width):
    return [stringify_cell(row[index] if index < len(row) else "") for index in range(width)]


def add_row_numbers_to_preview(columns, rows):
    preview_columns = [UPLOAD_ROW_NO_COLUMN, *columns]
    preview_rows = [
        [row_number, *row]
        for row_number, row in enumerate(rows or [], start=1)
    ]
    return preview_columns, preview_rows


def require_project_access(request, project_id="", project_code=""):
    """Verify the selected project without using browser-provided identity values."""
    request_user_id = get_request_user_id(request)
    normalized_project_id = str(project_id or "").strip()
    normalized_project_code = str(project_code or "").strip()
    if not normalized_project_id:
        return

    target_conn = None
    try:
        target_conn = get_target_db_connection(request)
        project_result = execute_query(target_conn, "M02001_PROJECT_OWNER_CONTEXT", {
            "projectId": normalized_project_id,
            "projectCode": normalized_project_code,
        })
        project_rows = project_result.get("data", []) if project_result.get("status") == "success" else []
        if not project_rows:
            raise HTTPException(status_code=404, detail="Project not found.")

        project_owner_user_id = int(project_rows[0].get("USER_ID") or 0)
        if project_owner_user_id <= 0:
            raise HTTPException(status_code=409, detail="Project owner is not configured.")
        if project_owner_user_id != int(request_user_id) and get_request_role_code(request) != "ADMIN":
            raise HTTPException(status_code=403, detail="Project access denied.")
    finally:
        if target_conn:
            target_conn.close()


def create_upload_table_name(project_code="", table_name_rule="", login_id="", user_id=""):
    timestamp = str(int(time.time() * 1000))
    project_token = normalize_identifier_token(project_code or "PROJECT")[:40] or "PROJECT"
    rule = (table_name_rule or "INITUP$_{PROJECT_CODE}_FT_{TIME}").strip()
    if "{TIME}" not in rule.upper():
        rule = f"{rule}_{{TIME}}"
    name = rule
    name = name.replace("{PROJECT_CODE}", project_token)
    name = name.replace("{project_code}", project_token)
    name = name.replace("{TIME}", timestamp)
    name = name.replace("{time}", timestamp)
    name = normalize_identifier_token(name)
    if not name.startswith("INITUP$_"):
        name = f"INITUP$_{name}"
    if not re.search(r"[0-9]{13}$", name):
        name = f"{name}_{timestamp}"
    if len(name) > 120:
        name = f"{name[:106].rstrip('_')}_{timestamp}"
    return name


def create_upload_table_prefix(project_code="", login_id=""):
    project_token = normalize_identifier_token(project_code or "PROJECT")[:40] or "PROJECT"
    return f"INITUP$_{project_token}_FT_"


def normalize_upload_table_search_prefix(table_prefix="", base_prefix=""):
    requested = normalize_upload_prefix_token(table_prefix or "")
    base = normalize_identifier_token(base_prefix or "INITUP$_PROJECT_FT_")
    if not requested:
        return base
    if not requested.startswith("INITUP$"):
        return base
    return requested


def normalize_upload_prefix_token(value):
    name = re.sub(r"[^A-Za-z0-9_$#]", "_", str(value or "").upper())
    return re.sub(r"_+", "_", name).lstrip("_")


def normalize_identifier_token(value):
    name = re.sub(r"[^A-Za-z0-9_$#]", "_", str(value or "").upper())
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def build_file_upload_column_specs(columns, has_header):
    use_header = str(has_header or "Y").upper() == "Y"
    return [
        (
            f"COL{index + 1:03d}",
            str(original_column or "").strip() if use_header else "",
        )
        for index, original_column in enumerate(columns)
    ]


def validate_reload_column_layout(target_columns, source_width):
    target_column_names = [require_read_table(row.get("COLUMN_NAME")) for row in target_columns or []]
    if not target_column_names or target_column_names[0] != UPLOAD_ROW_NO_COLUMN:
        raise HTTPException(
            status_code=409,
            detail=f"선택한 테이블의 첫 번째 컬럼은 {UPLOAD_ROW_NO_COLUMN}여야 합니다. 데이터는 변경되지 않았습니다.",
        )
    target_width = len(target_column_names) - 1
    if target_width != int(source_width or 0):
        raise HTTPException(
            status_code=409,
            detail=(
                "파일 컬럼 수가 선택한 테이블과 일치하지 않습니다. "
                f"파일: {int(source_width or 0)}개, 테이블(FILE_ROW_NO 제외): {target_width}개. "
                "데이터는 변경되지 않았습니다."
            ),
        )
    return target_column_names


def stringify_cell(value):
    if value is None:
        return ""
    return str(value)


def gather_upload_table_stats(cursor, table_name):
    cursor.execute(
        SqlLoader.get_sql("M02001_UPLOAD_TABLE_STATS_GATHER"),
        {"tableName": table_name},
    )


def require_upload_table(table_name):
    name = (table_name or "").strip().upper()
    if not name.startswith("INITUP$_"):
        raise HTTPException(status_code=400, detail="Only INITUP$_ tables can be used here.")
    if not re.fullmatch(r"INITUP\$_[A-Z0-9_$#_]*[0-9]{13}", name):
        raise HTTPException(status_code=400, detail="Invalid upload table ID.")
    return name


def require_read_table(table_name):
    name = (table_name or "").strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z0-9_$#]{0,127}", name):
        raise HTTPException(status_code=400, detail="Invalid table ID.")
    return name


def escape_and_truncate_oracle_comment(value, max_bytes=ORACLE_COMMENT_SAFE_BYTE_LIMIT):
    result = []
    used_bytes = 0
    for character in str(value or ""):
        escaped_character = "''" if character == "'" else character
        character_bytes = len(escaped_character.encode("utf-8"))
        if used_bytes + character_bytes > max_bytes:
            break
        result.append(escaped_character)
        used_bytes += character_bytes
    return "".join(result)


def normalize_limit(value):
    try:
        limit = int(value or 100)
    except (TypeError, ValueError):
        limit = 100
    return max(1, min(limit, 1000))


def normalize_select_sql(sql):
    text = (sql or "").strip()
    text = re.sub(r";+\s*$", "", text)
    if not re.match(r"(?is)^(select|with)\b", text):
        raise HTTPException(status_code=400, detail="Only SELECT statements are allowed.")
    if re.search(r";\s*\S", sql or ""):
        raise HTTPException(status_code=400, detail="Only a single SELECT statement is allowed.")
    blocked = r"\b(insert|update|delete|merge|drop|alter|create|truncate|grant|revoke|begin|declare|execute|exec)\b"
    if re.search(blocked, text, re.IGNORECASE):
        raise HTTPException(status_code=400, detail="Only read-only SELECT statements are allowed.")
    return text
